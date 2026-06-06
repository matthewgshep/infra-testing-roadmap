#!/usr/bin/env python3
"""
generate_gantt.py — Reads the Infra Testing Excel file and generates
an interactive HTML Gantt view, with optional deploy to GitHub Pages.

SETUP:
  pip install openpyxl

USAGE:
  python generate_gantt.py                   # generate HTML only
  python generate_gantt.py --deploy          # generate + push to GitHub Pages

FIRST-TIME GITHUB PAGES SETUP:
  1. Create a repo on your GitHub Enterprise (or github.com):
       gh repo create infra-testing-roadmap --public
  2. Clone it locally:
       git clone https://github.com/YOUR_ORG/infra-testing-roadmap.git
  3. Set the repo path below (GITHUB_REPO_PATH) or pass --repo
  4. Enable GitHub Pages: repo Settings > Pages > Source: main branch
  5. Run: python generate_gantt.py --deploy
  6. Your Gantt is live at: https://YOUR_ORG.github.io/infra-testing-roadmap/
"""

import argparse
import json
import os
import shutil
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

# ── Adobe fiscal year start dates (Saturday nearest Dec 1) ──────────────
# Auto-extends 3 fiscal years past today so new quarters appear automatically.
FY_STARTS = {
    2023: datetime(2022, 12, 3),
    2024: datetime(2023, 12, 2),
    2025: datetime(2024, 11, 30),
    2026: datetime(2025, 11, 29),
    2027: datetime(2026, 11, 28),
}

def _extend_fy_starts():
    """Auto-extend FY_STARTS to cover at least 3 fiscal years past today using a
    52-week (364-day) rolling extension. Adobe occasionally has 53-week fiscal
    years for calendar alignment; revisit and correct manually every ~5 years."""
    target_fy = datetime.now().year + 3
    while max(FY_STARTS.keys()) < target_fy:
        last_fy = max(FY_STARTS.keys())
        FY_STARTS[last_fy + 1] = FY_STARTS[last_fy] + timedelta(days=364)

_extend_fy_starts()


def fiscal_info(dt):
    for fy in sorted(FY_STARTS.keys(), reverse=True):
        fy_start = FY_STARTS[fy]
        if dt >= fy_start:
            days_in = (dt - fy_start).days
            q = min(days_in // 91 + 1, 4)
            q_start = fy_start + timedelta(days=91 * (q - 1))
            wk = min((dt - q_start).days // 7 + 1, 13)
            return fy, q, wk
    return 2023, 1, 1


def extract_data(excel_path):
    print(f"  Reading: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb['Results']

    # Also load with formulas to get hyperlinks
    wb_links = load_workbook(excel_path)
    ws_links = wb_links['Results']

    tests = []
    for r in range(2, ws.max_row + 1):
        product = ws.cell(r, 2).value
        if isinstance(product, str):
            product = product.replace('Acrobat Reader', 'Reader').replace('Acrobat Reduced Mode', 'Reduced Mode')
        name = ws.cell(r, 3).value
        start = ws.cell(r, 4).value
        end = ws.cell(r, 5).value
        qgnarr = ws.cell(r, 6).value
        gnarr = ws.cell(r, 7).value
        ctr = ws.cell(r, 8).value
        units_pct = ws.cell(r, 9).value
        winner = ws.cell(r, 10).value
        # Get GDS link text + hyperlink URL
        gds_text = ws.cell(r, 11).value
        gds_cell = ws_links.cell(r, 11)
        gds_url = gds_cell.hyperlink.target if gds_cell.hyperlink else None
        details = ws.cell(r, 12).value          # col 12: Decision/Details
        result_contrary = ws.cell(r, 13).value  # col 13: Result contrary to expectation
        pm_commentary = ws.cell(r, 14).value     # col 14: PM Questions/Commentary
        if not product or not isinstance(start, datetime):
            continue
        has_no_end = not isinstance(end, datetime)
        if has_no_end:
            end = datetime.today()
        is_scheduled = has_no_end and start > datetime.today()
        is_running = has_no_end and not is_scheduled
        fy, q, _ = fiscal_info(start)
        fq = f"FY{fy % 100:02d} Q{q}"

        def fmt_num(v):
            return round(v, 4) if isinstance(v, (int, float)) else None

        def fmt_str(v):
            s = str(v).strip() if v else None
            return s if s and s not in ('-', 'None', '') else None

        status = 'Scheduled' if is_scheduled else ('Running' if is_running else 'Complete')
        tests.append({
            'product': product,
            'name': name,
            'start': start.strftime('%Y-%m-%d'),
            'end': end.strftime('%Y-%m-%d'),
            'fq': fq,
            'qgnarr': round(qgnarr) if isinstance(qgnarr, (int, float)) else None,
            'gnarr': fmt_num(gnarr),
            'ctr': fmt_num(ctr),
            'units_pct': fmt_num(units_pct),
            'winner': fmt_str(winner),
            'status': status,
            'gds_text': fmt_str(gds_text),
            'gds_url': gds_url,
            'details': fmt_str(details),
            'result_contrary': fmt_str(result_contrary),
            'pm_commentary': fmt_str(pm_commentary),
        })

    # Generate fiscal quarter boundaries
    fq_bounds = []
    for fy in sorted(FY_STARTS.keys()):
        for q in range(1, 5):
            qs = FY_STARTS[fy] + timedelta(days=91 * (q - 1))
            qe = FY_STARTS[fy] + timedelta(days=91 * q - 1)
            fq_bounds.append({
                'label': f"FY{fy % 100:02d} Q{q}",
                'start': qs.strftime('%Y-%m-%d'),
                'end': qe.strftime('%Y-%m-%d'),
            })

    # Read backlog sheet
    backlog = []
    if 'Backlog' in wb.sheetnames:
        ws_bl = wb['Backlog']
        for r in range(2, ws_bl.max_row + 1):
            product = ws_bl.cell(r, 1).value
            name = ws_bl.cell(r, 2).value
            if product and name:
                backlog.append({'product': str(product).strip(), 'name': str(name).strip()})
        print(f"  Found {len(backlog)} backlog items")

    # Calculate average test duration per product (in weeks)
    avg_duration = {}
    for t in tests:
        if t['status'] == 'Complete':
            days = (datetime.strptime(t['end'], '%Y-%m-%d') - datetime.strptime(t['start'], '%Y-%m-%d')).days
            weeks = max(1, round(days / 7))
            avg_duration.setdefault(t['product'], []).append(weeks)
    avg_weeks = {p: round(sum(w) / len(w)) for p, w in avg_duration.items()}
    print(f"  Avg duration: {avg_weeks}")

    n_run = sum(1 for t in tests if t['status'] == 'Running')
    n_sched = sum(1 for t in tests if t['status'] == 'Scheduled')
    print(f"  Found {len(tests)} tests ({n_run} running, {n_sched} scheduled)")
    return {'tests': tests, 'quarters': fq_bounds, 'backlog': backlog, 'avgDuration': avg_weeks}


def generate_html(data):
    data_json = json.dumps(data)
    generated_ts = datetime.now().strftime('%b %d, %Y at %I:%M %p')

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Product Testing Roadmap — Gantt</title>
<style>
  :root {{
    --bg: #f8f9fa; --surface: #ffffff; --border: #e2e8f0;
    --text: #1a202c; --text-muted: #718096; --text-light: #a0aec0;
    --blue: #2E75B6; --blue-dark: #1F4E78; --blue-light: #D6E4F0;
    --green: #006100; --green-bg: #C6EFCE;
    --red: #9C0006; --red-bg: #FFC7CE;
    --gold: #FFC000; --gold-dark: #BF8F00; --gold-bg: #FFF2CC;
    --row-alt: #f7fafc;
  }}
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  html, body {{ height: 100%; overflow: hidden; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: var(--bg); color: var(--text); display: flex; flex-direction: column; }}
  .header {{ background: var(--blue-dark); color: white; padding: 20px 28px; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 16px; }}
  .header h1 {{ font-size: 20px; font-weight: 600; letter-spacing: -0.3px; }}
  .header .generated {{ font-size: 11px; opacity: 0.6; }}
  .controls {{ display: flex; gap: 12px; align-items: center; flex-wrap: wrap; }}
  .controls label {{ font-size: 13px; font-weight: 500; opacity: 0.85; }}
  .controls select, .controls input {{ padding: 6px 10px; border-radius: 6px; border: 1px solid rgba(255,255,255,0.25); background: rgba(255,255,255,0.12); color: white; font-size: 13px; outline: none; }}
  .controls select option {{ background: var(--blue-dark); color: white; }}
  .controls select:focus, .controls input:focus {{ border-color: rgba(255,255,255,0.6); }}
  .legend {{ display: flex; gap: 16px; padding: 10px 28px; background: var(--surface); border-bottom: 1px solid var(--border); font-size: 12px; align-items: center; }}
  .legend-item {{ display: flex; align-items: center; gap: 5px; }}
  .legend-swatch {{ width: 14px; height: 14px; border-radius: 3px; }}
  .stats {{ display: flex; gap: 20px; padding: 8px 28px; background: var(--blue-light); font-size: 12px; color: var(--blue-dark); font-weight: 500; }}
  .gantt-wrapper {{ display: flex; overflow: hidden; border-top: 1px solid var(--border); flex: 1; min-height: 0; }}
  .left-panel {{ flex-shrink: 0; overflow-y: auto; border-right: 2px solid var(--border); background: var(--surface); height: 100%; }}
  .left-panel::-webkit-scrollbar {{ display: none; }}
  .left-header {{ display: grid; grid-template-columns: 90px 1fr 62px 58px 52px 52px 75px; position: sticky; top: 0; z-index: 10; }}
  .left-header > div {{ padding: 6px 6px; font-size: 10px; font-weight: 600; color: white; background: var(--blue-dark); border-right: 1px solid rgba(255,255,255,0.1); white-space: nowrap; display:flex; align-items:center; }}
  .left-header .col-r {{ justify-content: flex-end; }}
  .left-row {{ display: grid; grid-template-columns: 90px 1fr 62px 58px 52px 52px 75px; border-bottom: 1px solid var(--border); cursor: pointer; transition: background 0.1s; }}
  .left-row:hover {{ background: var(--blue-light) !important; }}
  .left-row > div {{ padding: 4px 6px; font-size: 11px; overflow: hidden; text-overflow: ellipsis; border-right: 1px solid var(--border); display: flex; flex-direction: column; justify-content: center; }}
  .left-row .cell-main {{ white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
  .left-row .mcol {{ text-align: right; font-variant-numeric: tabular-nums; font-weight: 600; font-size: 10px; white-space: nowrap; cursor: default; padding: 4px 6px; position: relative; }}
  .left-row .mcol:hover {{ background: var(--blue-light); }}
  .left-row .mcol.pos {{ color: var(--green); }}
  .left-row .mcol.neg {{ color: var(--red); }}
  .left-row .mcol.neutral {{ color: var(--text-muted); }}
  #mcTip {{ position: fixed; background: var(--blue-dark); color: white; padding: 5px 10px; border-radius: 5px; font-size: 10px; white-space: nowrap; z-index: 9999; pointer-events: none; box-shadow: 0 3px 10px rgba(0,0,0,0.3); display: none; }}
  #mcTip::after {{ content: ''; position: absolute; top: 100%; left: 50%; transform: translateX(-50%); border: 5px solid transparent; border-top-color: var(--blue-dark); }}
  .left-row .winner-challenger {{ color: var(--green); background: var(--green-bg); font-weight: 600; border-radius: 3px; padding: 1px 6px; font-size: 10px; white-space: nowrap; }}
  .left-row .winner-control {{ color: var(--red); background: var(--red-bg); font-weight: 600; border-radius: 3px; padding: 1px 6px; font-size: 10px; white-space: nowrap; }}
  .left-row .winner-rollout {{ color: var(--blue-dark); background: var(--blue-light); font-weight: 600; border-radius: 3px; padding: 1px 6px; font-size: 10px; white-space: nowrap; }}
  .status-running {{ color: var(--gold-dark); background: var(--gold-bg); font-weight: 600; border-radius: 3px; padding: 1px 6px; font-size: 10px; white-space: nowrap; }}
  .status-scheduled {{ color: #6b21a8; background: #f3e8ff; font-weight: 600; border-radius: 3px; padding: 1px 6px; font-size: 10px; white-space: nowrap; }}
  .right-panel {{ flex: 1; overflow-x: auto; overflow-y: auto; height: 100%; }}
  .right-panel::-webkit-scrollbar {{ height: 10px; }}
  .right-panel::-webkit-scrollbar-thumb {{ background: #cbd5e0; border-radius: 5px; }}
  .timeline-header {{ position: sticky; top: 0; z-index: 9; }}
  .fq-row {{ display: flex; }}
  .fq-label {{ font-size: 11px; font-weight: 700; color: white; display: flex; align-items: center; justify-content: center; border-right: 2px solid rgba(255,255,255,0.3); }}
  .week-row {{ display: flex; }}
  .week-cell {{ font-size: 8px; color: rgba(255,255,255,0.8); text-align: center; padding: 2px 0; border-right: 1px solid rgba(255,255,255,0.15); display: flex; flex-direction: column; align-items: center; justify-content: center; line-height: 1.2; }}
  .week-cell .wk-num {{ font-weight: 700; font-size: 9px; color: white; }}
  .bar-row {{ display: flex; border-bottom: 1px solid var(--border); position: relative; }}
  .bar-cell {{ position: relative; border-right: 1px solid #f0f0f0; }}
  .bar-cell.qtr-start {{ border-left: 1.5px solid var(--blue-dark); }}
  .bar-fill {{ position: absolute; top: 3px; bottom: 3px; left: 0; right: 0; border-radius: 3px; }}
  .bar-fill.complete {{ background: var(--blue); }}
  .bar-fill.running {{ background: var(--gold); background: repeating-linear-gradient(45deg, var(--gold), var(--gold) 4px, #ffdb4d 4px, #ffdb4d 8px); }}
  .bar-fill.scheduled {{ background: repeating-linear-gradient(45deg, #a855f7, #a855f7 4px, #c084fc 4px, #c084fc 8px); opacity: 0.7; }}
  .tooltip {{ position: fixed; background: var(--surface); border: 1px solid var(--border); border-radius: 8px; padding: 14px 16px; font-size: 12px; box-shadow: 0 8px 24px rgba(0,0,0,0.15); pointer-events: none; z-index: 100; max-width: 380px; display: none; }}
  .tooltip .tt-title {{ font-weight: 700; font-size: 13px; margin-bottom: 6px; color: var(--blue-dark); }}
  .tooltip .tt-row {{ display: flex; justify-content: space-between; gap: 16px; padding: 2px 0; }}
  .tooltip .tt-label {{ color: var(--text-muted); }}
  .tooltip .tt-value {{ font-weight: 600; }}
  .tooltip hr {{ border: none; border-top: 1px solid var(--border); margin: 6px 0; }}
  .multi-select {{ position: relative; display: inline-block; }}
  .ms-display {{ padding: 6px 28px 6px 10px; border-radius: 6px; border: 1px solid rgba(255,255,255,0.25); background: rgba(255,255,255,0.12); color: white; font-size: 13px; cursor: pointer; min-width: 130px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 220px; }}
  .ms-display::after {{ content: '\\u25BC'; position: absolute; right: 8px; top: 50%; transform: translateY(-50%); font-size: 9px; opacity: 0.7; }}
  .ms-dropdown {{ display: none; position: absolute; top: 100%; left: 0; background: var(--blue-dark); border: 1px solid rgba(255,255,255,0.25); border-radius: 6px; z-index: 50; min-width: 160px; max-height: 280px; overflow-y: auto; margin-top: 4px; box-shadow: 0 8px 24px rgba(0,0,0,0.3); }}
  .ms-dropdown.open {{ display: block; }}
  .ms-dropdown label {{ display: flex; align-items: center; gap: 8px; padding: 6px 12px; font-size: 12px; color: white; cursor: pointer; white-space: nowrap; }}
  .ms-dropdown label:hover {{ background: rgba(255,255,255,0.1); }}
  .ms-dropdown input[type=checkbox] {{ accent-color: var(--gold); }}
  .ms-dropdown .ms-divider {{ border-top: 1px solid rgba(255,255,255,0.15); margin: 2px 0; }}
  .zoom-control {{ display: flex; align-items: center; gap: 6px; }}
  .zoom-control input[type=range] {{ width: 100px; accent-color: white; }}
  .detail-row {{ display: none; border-bottom: 1px solid var(--border); background: #f0f4f8; animation: slideDown 0.15s ease-out; }}
  .detail-row.open {{ display: flex; }}
  @keyframes slideDown {{ from {{ opacity: 0; max-height: 0; }} to {{ opacity: 1; max-height: 200px; }} }}
  .detail-left {{ width: 520px; flex-shrink: 0; padding: 10px 14px; border-right: 2px solid var(--border); }}
  .detail-right {{ flex: 1; }}
  .detail-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 4px 20px; font-size: 11px; }}
  .detail-grid .dl {{ color: var(--text-muted); font-size: 10px; text-transform: uppercase; letter-spacing: 0.3px; }}
  .detail-grid .dv {{ font-weight: 500; }}
  .metric-detail-row {{ display: flex; gap: 8px; margin: 8px 0; }}
  .metric-detail-card {{ flex: 1; background: white; border: 1px solid var(--border); border-radius: 6px; padding: 6px 8px; }}
  .metric-detail-card .mdc-top {{ display: flex; justify-content: space-between; align-items: baseline; margin-bottom: 4px; }}
  .metric-detail-card .mdc-label {{ font-size: 9px; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.3px; }}
  .metric-detail-card .mdc-val {{ font-size: 13px; font-weight: 700; }}
  .metric-detail-card .mdc-val.pos {{ color: var(--green); }}
  .metric-detail-card .mdc-val.neg {{ color: var(--red); }}
  .metric-detail-card .mdc-val.neutral {{ color: var(--text-muted); }}
  .metric-detail-card .mdc-bar {{ height: 6px; background: var(--border); border-radius: 3px; overflow: hidden; position: relative; }}
  .metric-detail-card .mdc-fill {{ position: absolute; top: 0; bottom: 0; border-radius: 0; }}
  .metric-detail-card .mdc-median {{ position: absolute; top: -2px; bottom: -2px; width: 2px; left: 50%; transform: translateX(-1px); background: var(--text-muted); z-index: 2; border-radius: 1px; }}
  .metric-detail-card .mdc-compare {{ font-size: 9px; color: var(--text-muted); margin-top: 3px; }}
  .detail-actions {{ display: flex; gap: 6px; margin-bottom: 8px; }}
  .detail-actions button {{ padding: 4px 10px; font-size: 10px; font-weight: 600; border: 1px solid var(--border); border-radius: 4px; background: white; color: var(--text); cursor: pointer; display: flex; align-items: center; gap: 4px; transition: all 0.15s; }}
  .detail-actions button:hover {{ background: var(--blue-light); border-color: var(--blue); color: var(--blue-dark); }}
  .detail-actions button.copied {{ background: var(--green-bg); border-color: var(--green); color: var(--green); }}
  .detail-notes {{ margin-top: 8px; font-size: 11px; line-height: 1.5; }}
  .detail-notes .dn-label {{ font-weight: 600; color: var(--blue-dark); font-size: 10px; text-transform: uppercase; letter-spacing: 0.3px; margin-bottom: 2px; }}
  .detail-notes .dn-text {{ color: var(--text); }}
  .left-row .expand-icon {{ font-size: 9px; color: var(--text-muted); margin-right: 4px; transition: transform 0.15s; }}
  .left-row.expanded .expand-icon {{ transform: rotate(90deg); }}
  .left-row.expanded {{ background: #e8f0fe !important; }}
  .revenue-strip {{ display: flex; gap: 0; padding: 0; background: var(--surface); border-bottom: 1px solid var(--border); overflow-x: auto; font-size: 11px; }}
  .revenue-strip:empty {{ display: none; }}
  .rev-card {{ flex: 0 0 auto; padding: 10px 18px; border-right: 1px solid var(--border); position: relative; }}
  .rev-card:last-child {{ border-right: none; }}
  .rev-card .rev-fq {{ font-weight: 700; color: var(--blue-dark); font-size: 12px; margin-bottom: 6px; }}
  .rev-card .rev-metrics {{ display: flex; gap: 16px; }}
  .rev-card .rev-item {{ display: flex; flex-direction: column; }}
  .rev-card .rev-label {{ font-size: 9px; text-transform: uppercase; letter-spacing: 0.4px; color: var(--text-muted); }}
  .rev-card .rev-val {{ font-size: 13px; font-weight: 700; color: var(--text); }}
  .rev-card .rev-val.highlight {{ color: var(--green); }}
  .rev-card.rev-total {{ background: var(--blue-light); }}
  .rev-card.rev-total .rev-fq {{ color: var(--blue-dark); }}
  .rev-expand {{ display: inline-flex; align-items: center; justify-content: center; width: 18px; height: 18px; border-radius: 50%; background: var(--blue-light); color: var(--blue-dark); font-size: 10px; cursor: pointer; margin-left: 6px; vertical-align: middle; transition: background 0.15s, transform 0.15s; }}
  .rev-expand:hover {{ background: var(--blue); color: white; }}
  .rev-expand.open {{ transform: rotate(180deg); }}
  .rev-detail-popout {{ display: none; position: fixed; z-index: 80; background: var(--surface); border: 1px solid var(--border); border-top: 2px solid var(--blue); border-radius: 0 0 8px 8px; box-shadow: 0 8px 24px rgba(0,0,0,0.15); max-height: 280px; overflow-y: auto; min-width: 500px; }}
  .rev-detail-popout.open {{ display: block; }}
  .rev-detail-table {{ width: 100%; border-collapse: collapse; font-size: 11px; }}
  .rev-detail-table th {{ text-align: left; padding: 6px 10px; font-size: 9px; text-transform: uppercase; letter-spacing: 0.4px; color: var(--text-muted); border-bottom: 1px solid var(--border); background: #f7fafc; position: sticky; top: 0; }}
  .rev-detail-table td {{ padding: 5px 10px; border-bottom: 1px solid #f0f0f0; }}
  .rev-detail-table tr:hover td {{ background: var(--blue-light); }}
  .rev-detail-table .rev-test-name {{ max-width: 200px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; font-weight: 500; }}
  .rev-detail-table .rev-num {{ text-align: right; font-weight: 600; font-variant-numeric: tabular-nums; }}
  .view-tabs {{ display: flex; gap: 0; padding: 0 28px; background: var(--surface); border-bottom: 1px solid var(--border); }}
  .view-tab {{ padding: 10px 20px; font-size: 13px; font-weight: 600; color: var(--text-muted); cursor: pointer; border-bottom: 3px solid transparent; transition: all 0.15s; }}
  .view-tab:hover {{ color: var(--blue-dark); }}
  .view-tab.active {{ color: var(--blue-dark); border-bottom-color: var(--blue); }}
  .view-panel {{ display: none; flex: 1; min-height: 0; }}
  .view-panel.active {{ display: flex; flex-direction: column; }}
  .list-container {{ flex: 1; overflow: auto; background: var(--surface); }}
  .exp-table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  .exp-table th {{ position: sticky; top: 0; z-index: 10; background: var(--blue-dark); color: white; text-align: left; padding: 8px 10px; font-size: 11px; font-weight: 600; letter-spacing: 0.3px; cursor: pointer; white-space: nowrap; user-select: none; border-right: 1px solid rgba(255,255,255,0.1); }}
  .exp-table th:hover {{ background: #163b5c; }}
  .exp-table th .sort-arrow {{ font-size: 9px; margin-left: 4px; opacity: 0.5; }}
  .exp-table th.sorted .sort-arrow {{ opacity: 1; }}
  .exp-table td {{ padding: 6px 10px; border-bottom: 1px solid var(--border); vertical-align: top; }}
  .exp-table tr:nth-child(even) td {{ background: var(--row-alt); }}
  .exp-table tr:hover td {{ background: var(--blue-light) !important; }}
  .exp-table .col-name {{ max-width: 280px; font-weight: 500; }}
  .exp-table .col-num {{ text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }}
  .exp-table .col-detail {{ max-width: 300px; font-size: 11px; color: var(--text-muted); line-height: 1.4; }}
  .exp-table .col-detail-text {{ display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical; overflow: hidden; }}
  .exp-table .col-detail-text.expanded {{ -webkit-line-clamp: unset; }}
  .exp-table a {{ color: var(--blue); text-decoration: none; }}
  .exp-table a:hover {{ text-decoration: underline; }}
  /* ── Planning Roadmap ── */
  .plan-wrapper {{ display: flex; flex: 1; min-height: 0; overflow: hidden; }}
  .plan-sidebar {{ width: 240px; flex-shrink: 0; background: var(--surface); border-right: 2px solid var(--border); display: flex; flex-direction: column; overflow: hidden; }}
  .plan-sidebar-header {{ padding: 12px 14px; font-size: 13px; font-weight: 700; color: var(--blue-dark); background: #f0f4f8; border-bottom: 1px solid var(--border); }}
  .plan-sidebar-list {{ flex: 1; overflow-y: auto; padding: 8px; }}
  .plan-backlog-card {{ padding: 8px 10px; margin-bottom: 6px; background: white; border: 1.5px solid var(--border); border-left: 3px solid var(--gold); border-radius: 6px; cursor: grab; font-size: 11px; transition: box-shadow 0.15s, border-color 0.15s; }}
  .plan-backlog-card:hover {{ border-color: var(--blue); box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
  .plan-backlog-card.dragging {{ opacity: 0.4; }}
  .plan-backlog-card .pbc-product {{ font-size: 9px; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.3px; }}
  .plan-backlog-card .pbc-name {{ font-weight: 600; margin-top: 2px; }}
  .plan-backlog-card.placed {{ display: none; }}
  .plan-main {{ flex: 1; display: flex; flex-direction: column; overflow: hidden; }}
  .plan-toolbar {{ display: flex; align-items: center; gap: 12px; padding: 8px 16px; background: #f0f4f8; border-bottom: 1px solid var(--border); font-size: 12px; }}
  .plan-toolbar button {{ padding: 5px 14px; border-radius: 5px; border: 1px solid var(--border); background: white; font-size: 11px; font-weight: 600; cursor: pointer; transition: all 0.15s; }}
  .plan-toolbar button:hover {{ background: var(--blue-light); border-color: var(--blue); }}
  .plan-toolbar button.primary {{ background: var(--blue); color: white; border-color: var(--blue); }}
  .plan-toolbar button.primary:hover {{ background: var(--blue-dark); }}
  .plan-timeline {{ flex: 1; overflow: auto; position: relative; }}
  .plan-lane {{ position: relative; border-bottom: 1px solid var(--border); }}
  .plan-lane:nth-child(even) {{ background: var(--row-alt); }}
  .plan-lane-label {{ position: absolute; left: 0; top: 0; bottom: 0; width: 180px; display: flex; align-items: center; padding: 0 10px; font-size: 11px; font-weight: 500; z-index: 2; background: inherit; border-right: 1px solid var(--border); }}
  .plan-bar {{ position: absolute; height: 24px; top: 50%; transform: translateY(-50%); border-radius: 4px; cursor: grab; font-size: 10px; font-weight: 600; color: white; display: flex; align-items: center; padding: 0 6px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; min-width: 20px; user-select: none; z-index: 3; }}
  .plan-bar.existing {{ background: var(--blue); opacity: 0.5; cursor: default; }}
  .plan-bar.running {{ background: repeating-linear-gradient(45deg, var(--gold), var(--gold) 4px, #ffdb4d 4px, #ffdb4d 8px); opacity: 0.5; cursor: default; color: var(--text); }}
  .plan-bar.scheduled-plan {{ background: repeating-linear-gradient(45deg, #a855f7, #a855f7 4px, #c084fc 4px, #c084fc 8px); opacity: 0.5; cursor: default; color: white; }}
  .plan-bar.planned {{ background: #8b5cf6; box-shadow: 0 2px 6px rgba(139,92,246,0.3); cursor: grab; }}
  .plan-bar.planned:hover {{ box-shadow: 0 3px 10px rgba(139,92,246,0.4); }}
  .plan-bar .plan-resize {{ position: absolute; top: 0; bottom: 0; width: 8px; cursor: ew-resize; z-index: 4; }}
  .plan-bar .plan-resize.left {{ left: -2px; }}
  .plan-bar .plan-resize.right {{ right: -2px; }}
  .plan-bar .plan-remove {{ position: absolute; top: -6px; right: -6px; width: 16px; height: 16px; border-radius: 50%; background: var(--red); color: white; font-size: 10px; display: none; align-items: center; justify-content: center; cursor: pointer; z-index: 5; }}
  .plan-bar.planned:hover .plan-remove {{ display: flex; }}
  .plan-ghost {{ position: fixed; padding: 6px 10px; background: #8b5cf6; color: white; border-radius: 5px; font-size: 11px; font-weight: 600; pointer-events: none; z-index: 100; opacity: 0.85; white-space: nowrap; }}
  .plan-drop-indicator {{ position: absolute; top: 0; bottom: 0; background: rgba(139,92,246,0.1); border: 2px dashed #8b5cf6; border-radius: 4px; pointer-events: none; z-index: 1; display: none; }}
  .scatter-container {{ flex: 1; position: relative; padding: 0; overflow: hidden; background: var(--surface); }}
  .scatter-svg {{ width: 100%; height: 100%; }}
  .scatter-dot {{ cursor: pointer; transition: r 0.1s; }}
  .scatter-dot:hover {{ r: 7; }}
  .scatter-label {{ font-size: 9px; font-weight: 600; pointer-events: none; }}
  .scatter-avg-line {{ stroke-dasharray: 6 4; }}

  .surface-toggle {{ display: flex; gap: 2px; background: var(--border); border-radius: 6px; padding: 2px; }}
  .surface-btn {{ padding: 4px 10px; border: none; background: transparent; border-radius: 4px; font-size: 11px; font-weight: 600; color: var(--text-muted); cursor: pointer; transition: all 0.15s; white-space: nowrap; }}
  .surface-btn:hover {{ background: rgba(255,255,255,0.6); color: var(--text); }}
  .surface-btn.active {{ background: white; color: var(--blue-dark); box-shadow: 0 1px 3px rgba(0,0,0,0.12); }}
  .surface-btn[data-val="Reader"].active {{ color: #1d4ed8; }}
  .surface-btn[data-val="Reduced Mode"].active {{ color: #7c3aed; }}

</style>
</head>
<body>
<div class="header">
  <div><h1>Product Testing Roadmap</h1><div class="generated">Generated {generated_ts}</div></div>
  <div class="controls">
    <label>Quarter</label>
    <div class="multi-select" id="fqFilter">
      <div class="ms-display" id="fqDisplay">All Quarters</div>
      <div class="ms-dropdown" id="fqDropdown"></div>
    </div>
    <label>Surface</label>
    <div class="surface-toggle" id="surfaceToggle">
      <button class="surface-btn active" data-val="All">Both</button>
      <button class="surface-btn" data-val="Reader">Reader</button>
      <button class="surface-btn" data-val="Reduced Mode">Reduced Mode</button>
    </div>
    <select id="productFilter" style="display:none;"><option value="All">All Products</option></select>
    <label>Status</label>
    <select id="statusFilter"><option value="All">All</option><option value="Complete">Complete</option><option value="Running">Running</option><option value="Scheduled">Scheduled</option></select>
    <div class="zoom-control"><label>Zoom</label><input type="range" id="zoomSlider" min="8" max="80" value="80"></div>
    <label>Search</label>
    <input type="text" id="searchInput" placeholder="Test name..." style="width:150px;">
  </div>
</div>
<div class="view-tabs">
  <div class="view-tab active" data-view="gantt">Results Roadmap</div>
  <div class="view-tab" data-view="scatter">Experiment Lift</div>
  <div class="view-tab" data-view="list">Experiment List</div>
  <div class="view-tab" data-view="plan">Planning Roadmap</div>
  <div class="view-tab" data-view="velocity">Velocity</div>
</div>
<div class="view-panel active" id="ganttView">
<div class="legend" style="display:none;">
</div>
<div class="stats" id="statsBar"></div>
<div class="revenue-strip" id="revenueStrip"></div>
<div class="gantt-wrapper" id="ganttWrapper">
  <div class="left-panel" id="leftPanel"></div>
  <div class="right-panel" id="rightPanel"></div>
</div>
</div>
<div class="view-panel" id="scatterView">
  <div style="padding:6px 12px;display:flex;align-items:center;gap:8px;background:var(--surface);border-bottom:1px solid var(--border);">
    <label style="font-size:11px;font-weight:600;color:var(--text-muted);">Metric</label>
    <select id="scatterMetric" style="font-size:11px;padding:3px 8px;border:1px solid var(--border);border-radius:4px;background:white;">
      <option value="gnarr">GNARR Lift</option>
      <option value="ctr">CTR</option>
      <option value="units_pct">Units %</option>
    </select>
  </div>
  <div class="scatter-container" id="scatterContainer">
    <svg class="scatter-svg" id="scatterSvg"></svg>
  </div>
</div>
<div class="view-panel" id="listView">
  <div class="list-container" id="listContainer"></div>
</div>
<div class="view-panel" id="planView">
  <div class="plan-wrapper">
    <div class="plan-sidebar">
      <div class="plan-sidebar-header">Backlog <span id="planBacklogCount"></span></div>
      <div class="plan-sidebar-list" id="planBacklogList"></div>
    </div>
    <div class="plan-main">
      <div class="plan-toolbar">
        <span style="font-weight:600;color:var(--blue-dark);">Drag experiments from the backlog onto the timeline</span>
        <span id="planSaveStatus" style="font-size:10px;color:var(--text-muted);"></span>
        <span style="flex:1;"></span>
        <button onclick="planClearAll()">Clear All</button>
        <button onclick="planExport()">Export JSON</button>
        <button class="primary" id="planSaveBtn" disabled style="opacity:0.5;" onclick="planSaveToGitHub()">Save Plan</button>
        <button onclick="ghClearToken()" style="padding:5px 8px;font-size:9px;color:var(--text-muted);border:none;background:none;cursor:pointer;" title="Clear saved GitHub token">\u2699</button>
      </div>
      <div class="plan-timeline" id="planTimeline"></div>
    </div>
  </div>
</div>
<div class="view-panel" id="velocityView">
  <div class="scatter-container" id="velocityContainer">
    <canvas id="velocityCanvas" style="width:100%;height:100%;"></canvas>
  </div>
</div>
<div class="tooltip" id="tooltip"></div>
<div id="mcTip"></div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/chartjs-plugin-annotation/3.0.1/chartjs-plugin-annotation.min.js"></script>
<script>
const RAW_DATA = {data_json};
const TESTS = RAW_DATA.tests;
const ALL_QUARTERS = RAW_DATA.quarters;
const QUARTERS = ALL_QUARTERS.filter(q => {{
  const qStart = new Date(q.start), qEnd = new Date(q.end);
  // Always include the current quarter even if it has no tests yet
  const today = new Date();
  if (today >= qStart && today <= qEnd) return true;
  return TESTS.some(t => {{
    const ts = new Date(t.start), te = new Date(t.end);
    return ts <= qEnd && te >= qStart;
  }});
}});
const WEEKS = [];
QUARTERS.forEach(q => {{
  const qs = new Date(q.start);
  for (let w = 0; w < 13; w++) {{
    const ws = new Date(qs.getTime() + w * 7 * 86400000);
    const we = new Date(ws.getTime() + 6 * 86400000);
    WEEKS.push({{ start: ws, end: we, num: w + 1, fq: q.label }});
  }}
}});
// Quarters offered in the filter: any quarter that has tests, plus the current and
// next quarter — so as one quarter ends, the next is already available for tests that
// run into it. Ordered newest-first by the fiscal calendar.
const _now = new Date();
const _curQ = ALL_QUARTERS.find(q => _now >= new Date(q.start) && _now <= new Date(q.end));
const _curIdx = _curQ ? ALL_QUARTERS.indexOf(_curQ) : -1;
const _offeredFQ = new Set(TESTS.map(t => t.fq));
if (_curQ) _offeredFQ.add(_curQ.label);
if (_curIdx >= 0 && ALL_QUARTERS[_curIdx + 1]) _offeredFQ.add(ALL_QUARTERS[_curIdx + 1].label);
const fqSet = ALL_QUARTERS.filter(q => _offeredFQ.has(q.label)).map(q => q.label).reverse();
const productSet = [...new Set(TESTS.map(t => t.product))];
const prodSel = document.getElementById('productFilter');
productSet.forEach(p => {{ const o = document.createElement('option'); o.value = p; o.text = p; prodSel.add(o); }});
let activeSurface = 'All';
// Multi-select quarter filter
const fqDropdown = document.getElementById('fqDropdown');
const fqDisplay = document.getElementById('fqDisplay');
// Default the filter to the current quarter (fall back to the newest quarter with tests)
const latestFQ = (_curQ && fqSet.includes(_curQ.label)) ? _curQ.label : fqSet[0];
let selectedFQs = new Set([latestFQ]);
const allLabel = document.createElement('label');
allLabel.innerHTML = `<input type="checkbox" id="fqAll"> All Quarters`;
fqDropdown.appendChild(allLabel);
const divider = document.createElement('div');
divider.className = 'ms-divider';
fqDropdown.appendChild(divider);
fqSet.forEach(f => {{
  const label = document.createElement('label');
  const checked = f === latestFQ ? 'checked' : '';
  label.innerHTML = `<input type="checkbox" class="fq-cb" value="${{f}}" ${{checked}}> ${{f}}`;
  fqDropdown.appendChild(label);
}});
fqDisplay.textContent = latestFQ;
document.getElementById('fqAll').addEventListener('change', (e) => {{
  if (e.target.checked) {{
    selectedFQs.clear();
    document.querySelectorAll('.fq-cb').forEach(cb => cb.checked = false);
    fqDisplay.textContent = 'All Quarters';
    renderActive();
  }}
}});
document.querySelectorAll('.fq-cb').forEach(cb => {{
  cb.addEventListener('change', () => {{
    if (cb.checked) {{
      selectedFQs.add(cb.value);
      document.getElementById('fqAll').checked = false;
    }} else {{
      selectedFQs.delete(cb.value);
      if (selectedFQs.size === 0) document.getElementById('fqAll').checked = true;
    }}
    fqDisplay.textContent = selectedFQs.size === 0 ? 'All Quarters' : [...selectedFQs].join(', ');
    renderActive();
  }});
}});
fqDisplay.addEventListener('click', (e) => {{
  e.stopPropagation();
  fqDropdown.classList.toggle('open');
}});
document.addEventListener('click', () => fqDropdown.classList.remove('open'));
fqDropdown.addEventListener('click', (e) => e.stopPropagation());
// Initial CELL_W — will be recalculated after first render to fit 13 weeks
let CELL_W = 22;
const ROW_H = 32;
const tooltip = document.getElementById('tooltip');
function getFiltered() {{
  const prod = activeSurface;
  const status = document.getElementById('statusFilter').value;
  const search = document.getElementById('searchInput').value.toLowerCase();
  return TESTS.filter(t => {{
    if (selectedFQs.size > 0) {{
      // Show test if it overlaps with any selected quarter
      const ts = new Date(t.start), te = new Date(t.end);
      const overlaps = QUARTERS.some(q => {{
        if (!selectedFQs.has(q.label)) return false;
        const qs = new Date(q.start), qe = new Date(q.end);
        return ts <= qe && te >= qs;
      }});
      if (!overlaps) return false;
    }}
    if (prod !== 'All' && t.product !== prod) return false;
    if (status !== 'All' && t.status !== status) return false;
    if (search && !t.name.toLowerCase().includes(search)) return false;
    return true;
  }});
}}
function getVisibleWeeks() {{
  if (selectedFQs.size === 0) return WEEKS;
  return WEEKS.filter(w => selectedFQs.has(w.fq));
}}
function fmtDate(s) {{ const d = new Date(s); return d.toLocaleDateString('en-US', {{ month: 'short', day: 'numeric', year: 'numeric' }}); }}
function fmtPct(v) {{ return v !== null ? (v >= 0 ? '+' : '') + (v * 100).toFixed(2) + '%' : '\\u2014'; }}
function esc(s) {{ return s ? s.replace(/</g,'&lt;').replace(/>/g,'&gt;') : ''; }}
function fmtDollar(v) {{ return v !== null ? '$' + v.toLocaleString() : '\\u2014'; }}
function metricColor(v) {{ return v > 0 ? 'var(--green)' : v < 0 ? 'var(--red)' : 'inherit'; }}
function fmtMetric(v) {{
  if (v === null) return '\\u2014';
  const pct = (v * 100).toFixed(2);
  const sign = v > 0 ? '+' : '';
  const color = metricColor(v);
  return `<span style="color:${{color}};font-weight:600;">${{sign}}${{pct}}%</span>`;
}}
function buildTooltipHtml(t) {{
  const dur = Math.round((new Date(t.end) - new Date(t.start)) / 86400000);
  const statusCls = t.status === 'Running' ? 'status-running' : t.status === 'Scheduled' ? 'status-scheduled' : '';
  const statusHtml = statusCls ? `<span class="${{statusCls}}">${{t.status}}</span>` : t.status;
  let winHtml = '\\u2014';
  if (t.winner) {{
    const wCls = t.winner.startsWith('Challenger') ? 'winner-challenger' : t.winner === 'Control' ? 'winner-control' : '';
    winHtml = wCls ? `<span class="${{wCls}}">${{t.winner}}</span>` : esc(t.winner);
  }}
  return `
    <div class="tt-title">${{esc(t.name)}}</div>
    <div class="tt-row"><span class="tt-label">Product</span><span class="tt-value">${{esc(t.product)}}</span></div>
    <div class="tt-row"><span class="tt-label">Dates</span><span class="tt-value">${{fmtDate(t.start)}} \\u2014 ${{(t.status === 'Running' || t.status === 'Scheduled') ? 'TBD' : fmtDate(t.end)}}</span></div>
    <div class="tt-row"><span class="tt-label">Duration</span><span class="tt-value">${{dur}} days</span></div>
    <div class="tt-row"><span class="tt-label">Status</span><span class="tt-value">${{statusHtml}}</span></div>
    <div class="tt-row"><span class="tt-label">Winner</span><span class="tt-value">${{winHtml}}</span></div>
    <hr>
    <div class="tt-row"><span class="tt-label">QGNARR</span><span class="tt-value">${{fmtDollar(t.qgnarr)}}</span></div>
    <div class="tt-row"><span class="tt-label">GNARR Lift</span><span class="tt-value">${{fmtMetric(t.gnarr)}}</span></div>
    <div class="tt-row"><span class="tt-label">CTR</span><span class="tt-value">${{fmtMetric(t.ctr)}}</span></div>
    <div class="tt-row"><span class="tt-label">Units %</span><span class="tt-value">${{fmtMetric(t.units_pct)}}</span></div>
  `;
}}
function fmtCompact(v) {{
  if (v === null || v === 0) return '$0';
  if (Math.abs(v) >= 1000000) return '$' + (v / 1000000).toFixed(1) + 'M';
  if (Math.abs(v) >= 1000) return '$' + (v / 1000).toFixed(0) + 'K';
  return '$' + Math.round(v).toLocaleString();
}}
function calcRevenueByQuarter(tests) {{
  const visibleFQs = selectedFQs.size > 0 ? [...selectedFQs] : [];
  if (visibleFQs.length === 0 || visibleFQs.length > 4) return [];
  const qMap = {{}};
  QUARTERS.forEach(q => {{
    if (!visibleFQs.includes(q.label)) return;
    qMap[q.label] = {{ label: q.label, start: new Date(q.start), end: new Date(q.end), qgnarrTotal: 0, realized: 0, projected: 0, testCount: 0, contributions: [] }};
  }});
  const DAY_MS = 86400000;
  const Q_DAYS = 91; // 13 weeks × 7 days
  const today = new Date();
  today.setHours(0, 0, 0, 0);
  tests.forEach(t => {{
    if (t.qgnarr === null || t.status === 'Running' || t.status === 'Scheduled') return;
    const ts = new Date(t.start), te = new Date(t.end);
    const dailyValue = t.qgnarr / Q_DAYS;
    const isWinner = t.winner && (t.winner === 'Rollout' || t.winner.startsWith('Challenger'));
    Object.values(qMap).forEach(qObj => {{
      const qs = qObj.start, qe = qObj.end;
      if (te >= qs && te <= qe) {{
        // Realized: days from test end to min(today, quarter end)
        const realizedEnd = today < qe ? today : qe;
        const realizedDays = Math.min(Q_DAYS, Math.max(0, Math.round((realizedEnd.getTime() - te.getTime()) / DAY_MS) + 1));
        const realizedAmt = dailyValue * realizedDays;
        // Remaining: days from today (or test end, whichever is later) to quarter end
        const remainStart = today > te ? today : te;
        const remainingDays = Math.max(0, Math.round((qe.getTime() - remainStart.getTime()) / DAY_MS));
        qObj.qgnarrTotal += t.qgnarr;
        qObj.realized += realizedAmt;
        qObj.testCount++;
        let remainingAmt = 0;
        if (isWinner && remainingDays > 0) {{
          remainingAmt = dailyValue * remainingDays;
          qObj.projected += remainingAmt;
        }}
        qObj.contributions.push({{
          name: t.name, product: t.product, winner: t.winner,
          qgnarr: t.qgnarr, realized: realizedAmt, remaining: remainingAmt,
          endDate: t.end, days: realizedDays
        }});
      }}
    }});
  }});
  return Object.values(qMap);
}}
function renderRevenueStrip(filtered) {{
  const strip = document.getElementById('revenueStrip');
  const revData = calcRevenueByQuarter(filtered);
  if (revData.length === 0) {{ strip.innerHTML = ''; return; }}
  let html = '';
  revData.forEach((q, qi) => {{
    // Build detail table rows
    const sorted = q.contributions.slice().sort((a, b) => b.qgnarr - a.qgnarr);
    let tableRows = '';
    if (sorted.length === 0) {{
      tableRows = `<tr><td colspan="6" style="text-align:center;color:var(--text-muted);padding:14px;">No completed tests in this quarter</td></tr>`;
    }}
    sorted.forEach(c => {{
      const winCls = c.winner && c.winner.startsWith('Challenger') ? 'winner-challenger' : c.winner === 'Control' ? 'winner-control' : c.winner === 'Rollout' ? 'winner-rollout' : '';
      const winHtml = c.winner ? `<span class="${{winCls}}">${{c.winner}}</span>` : '\\u2014';
      tableRows += `<tr>
        <td class="rev-test-name" title="${{esc(c.name)}}">${{esc(c.name)}}</td>
        <td>${{esc(c.product)}}</td>
        <td>${{winHtml}}</td>
        <td class="rev-num">${{fmtDollar(c.qgnarr)}}</td>
        <td class="rev-num" style="color:var(--green);">${{fmtCompact(c.realized + c.remaining)}}</td>
        <td class="rev-num">${{c.days}}d</td>
      </tr>`;
    }});
    const inQuarterTotal = q.realized + q.projected;
    html += `<div class="rev-card" style="position:relative;">
      <div class="rev-fq">${{q.label}}<span class="rev-expand" data-qi="${{qi}}">&#9660;</span></div>
      <div class="rev-metrics">
        <div class="rev-item"><span class="rev-label">QGNARR Projection</span><span class="rev-val">${{fmtCompact(q.qgnarrTotal)}}</span></div>
        <div class="rev-item"><span class="rev-label">In Quarter Realization Projection</span><span class="rev-val highlight">${{fmtCompact(inQuarterTotal)}}</span></div>
        <div class="rev-item"><span class="rev-label">Tests</span><span class="rev-val">${{q.testCount}}</span></div>
      </div>
      <div class="rev-detail-popout" id="revDetail${{qi}}">
        <table class="rev-detail-table">
          <thead><tr><th>Test Name</th><th>Product</th><th>Winner</th><th>QGNARR</th><th>Realized</th><th>Days</th></tr></thead>
          <tbody>${{tableRows}}</tbody>
        </table>
      </div>
    </div>`;
  }});
  strip.innerHTML = html;
  // Wire up expand buttons
  strip.querySelectorAll('.rev-expand').forEach(btn => {{
    btn.addEventListener('mousedown', (e) => {{
      e.preventDefault();
      e.stopPropagation();
      const qi = btn.dataset.qi;
      const popout = document.getElementById('revDetail' + qi);
      const isOpen = popout.classList.contains('open');
      // Close all first
      strip.querySelectorAll('.rev-detail-popout').forEach(p => p.classList.remove('open'));
      strip.querySelectorAll('.rev-expand').forEach(b => b.classList.remove('open'));
      if (!isOpen) {{
        // Position the popout below the card using fixed coords
        const card = btn.closest('.rev-card');
        const rect = card.getBoundingClientRect();
        popout.style.top = rect.bottom + 'px';
        popout.style.left = rect.left + 'px';
        popout.style.minWidth = Math.max(500, rect.width) + 'px';
        popout.classList.add('open');
        btn.classList.add('open');
      }}
    }});
  }});
  // Stop clicks inside the popout from closing it
  strip.querySelectorAll('.rev-detail-popout').forEach(p => {{
    p.addEventListener('mousedown', (e) => e.stopPropagation());
  }});
}}
// Global: close revenue popouts when clicking anywhere else
document.addEventListener('mousedown', () => {{
  const strip = document.getElementById('revenueStrip');
  if (strip) {{
    strip.querySelectorAll('.rev-detail-popout').forEach(p => p.classList.remove('open'));
    strip.querySelectorAll('.rev-expand').forEach(b => b.classList.remove('open'));
  }}
}});
// Compute per-product medians & maxes for metric detail cards
const _gMedians = {{}};
const _gMaxes = {{}};
const _gMetricKeys = ['qgnarr', 'gnarr', 'ctr', 'units_pct'];
TESTS.filter(t => t.status === 'Complete').forEach(t => {{
  if (!_gMedians[t.product]) {{ _gMedians[t.product] = {{}}; _gMaxes[t.product] = {{}}; }}
  _gMetricKeys.forEach(k => {{
    if (t[k] !== null) {{
      if (!_gMedians[t.product][k]) {{ _gMedians[t.product][k] = []; _gMaxes[t.product][k] = 0; }}
      _gMedians[t.product][k].push(t[k]);
      _gMaxes[t.product][k] = Math.max(_gMaxes[t.product][k], Math.abs(t[k]));
    }}
  }});
}});
function _gMedian(arr) {{
  if (!arr || !arr.length) return null;
  const s = arr.slice().sort((a, b) => a - b);
  const mid = Math.floor(s.length / 2);
  return s.length % 2 ? s[mid] : (s[mid - 1] + s[mid]) / 2;
}}
const gProductMedians = {{}};
const gProductMaxes = {{}};
Object.keys(_gMedians).forEach(prod => {{
  gProductMedians[prod] = {{}};
  gProductMaxes[prod] = {{}};
  _gMetricKeys.forEach(k => {{
    gProductMedians[prod][k] = _gMedian(_gMedians[prod][k]);
    gProductMaxes[prod][k] = _gMaxes[prod][k] || 1;
  }});
}});
function buildMetricCard(label, key, val, product, isCurrency) {{
  const pm = gProductMedians[product] || {{}};
  const med = pm[key];
  // No data
  if (val === null) return `<div class="metric-detail-card"><div class="mdc-top"><span class="mdc-label">${{label}}</span><span class="mdc-val neutral">\\u2014</span></div><div class="mdc-bar"><div class="mdc-median"></div></div><div class="mdc-compare">No data</div></div>`;
  const cls = val > 0 ? 'pos' : val < 0 ? 'neg' : 'neutral';
  const display = isCurrency ? fmtDollar(val) : fmtPct(val);
  // Diverging bar: median at center (50%), bar extends left or right
  let barHtml = '<div class="mdc-median"></div>';
  let compare = '';
  if (med !== null && med !== undefined) {{
    const diff = val - med;
    // Scale using 75th percentile of deviations so most bars are clearly visible
    const allVals = (_gMedians[product] && _gMedians[product][key]) || [];
    const devs = allVals.map(v => Math.abs(v - med)).sort((a, b) => a - b);
    const p75idx = Math.floor(devs.length * 0.75);
    const scaleDev = devs[p75idx] || devs[devs.length - 1] || 1;
    const pct = Math.min(50, Math.round((Math.abs(diff) / scaleDev) * 40));
    const barColor = diff > 0 ? 'var(--green)' : diff < 0 ? 'var(--red)' : 'var(--text-muted)';
    if (diff > 0) {{
      barHtml += `<div class="mdc-fill" style="left:50%;width:${{pct}}%;background:${{barColor}};"></div>`;
    }} else if (diff < 0) {{
      barHtml += `<div class="mdc-fill" style="right:50%;width:${{pct}}%;background:${{barColor}};"></div>`;
    }}
    if (med === 0) {{
      compare = isCurrency ? 'Median: $0' : 'Median: 0.00%';
    }} else {{
      const pctDiff = ((val - med) / Math.abs(med)) * 100;
      const sign = pctDiff > 0 ? '+' : '';
      const arrow = pctDiff > 0 ? '\\u25B2' : pctDiff < 0 ? '\\u25BC' : '';
      compare = `${{arrow}} ${{sign}}${{Math.round(pctDiff)}}% vs median (${{isCurrency ? fmtDollar(med) : fmtPct(med)}})`;
    }}
  }}
  return `<div class="metric-detail-card"><div class="mdc-top"><span class="mdc-label">${{label}}</span><span class="mdc-val ${{cls}}">${{display}}</span></div><div class="mdc-bar">${{barHtml}}</div><div class="mdc-compare">${{compare}}</div></div>`;
}}
function copyExperiment(idx) {{
  const t = TESTS[idx];
  if (!t) return;
  const dur = Math.round((new Date(t.end) - new Date(t.start)) / 86400000);
  let text = t.name + '\\n';
  text += 'Product: ' + t.product + '\\n';
  text += 'Fiscal Quarter: ' + t.fq + '\\n';
  text += 'Dates: ' + fmtDate(t.start) + ' \\u2014 ' + fmtDate(t.end) + ' (' + dur + ' days)\\n';
  text += 'Status: ' + t.status + '\\n';
  text += 'Winner: ' + (t.winner || '\\u2014') + '\\n';
  text += '\\nMetrics:\\n';
  text += '  QGNARR: ' + fmtDollar(t.qgnarr) + '\\n';
  text += '  GNARR Lift: ' + fmtPct(t.gnarr) + '\\n';
  text += '  CTR: ' + fmtPct(t.ctr) + '\\n';
  text += '  Units: ' + fmtPct(t.units_pct) + '\\n';
  if (t.details) text += '\\nDecision / Details:\\n' + t.details + '\\n';
  if (t.pm_commentary) text += '\\nPM Commentary:\\n' + t.pm_commentary + '\\n';
  if (t.gds_url) text += '\\nGDS: ' + t.gds_url + '\\n';
  navigator.clipboard.writeText(text).then(() => {{
    const btn = document.querySelector(`[data-copy-idx="${{idx}}"]`);
    if (btn) {{
      btn.classList.add('copied');
      btn.innerHTML = '\\u2713 Copied';
      setTimeout(() => {{ btn.classList.remove('copied'); btn.innerHTML = '\\u2398 Copy'; }}, 1500);
    }}
  }});
}}
function buildDetailHtml(t) {{
  const dur = Math.round((new Date(t.end) - new Date(t.start)) / 86400000);
  const tIdx = TESTS.indexOf(t);
  let html = '<div class="detail-actions">';
  html += `<button data-copy-idx="${{tIdx}}" onclick="event.stopPropagation();copyExperiment(${{tIdx}})">\\u2398 Copy</button>`;
  html += '</div>';
  html += '<div class="detail-grid">';
  html += `<div><span class="dl">Fiscal Quarter</span><div class="dv">${{t.fq}}</div></div>`;
  html += `<div><span class="dl">Duration</span><div class="dv">${{dur}} days</div></div>`;
  html += `<div><span class="dl">Dates</span><div class="dv">${{fmtDate(t.start)}} \\u2014 ${{fmtDate(t.end)}}</div></div>`;
  html += `<div><span class="dl">Status</span><div class="dv">${{t.status}}</div></div>`;
  html += `<div><span class="dl">Winner</span><div class="dv">${{t.winner || '\\u2014'}}</div></div>`;
  if (t.result_contrary) html += `<div><span class="dl">Result Contrary</span><div class="dv">${{esc(t.result_contrary)}}</div></div>`;
  html += '</div>';
  // Option C metric detail cards
  html += '<div class="metric-detail-row">';
  html += buildMetricCard('QGNARR', 'qgnarr', t.qgnarr, t.product, true);
  html += buildMetricCard('GNARR Lift', 'gnarr', t.gnarr, t.product, false);
  html += buildMetricCard('CTR', 'ctr', t.ctr, t.product, false);
  html += buildMetricCard('Units', 'units_pct', t.units_pct, t.product, false);
  html += '</div>';
  if (t.details) html += `<div class="detail-notes"><div class="dn-label">Decision / Details</div><div class="dn-text">${{esc(t.details)}}</div></div>`;
  if (t.pm_commentary) html += `<div class="detail-notes"><div class="dn-label">PM Commentary</div><div class="dn-text">${{esc(t.pm_commentary)}}</div></div>`;
  if (t.gds_text || t.gds_url) {{
    let gdsHtml = '';
    if (t.gds_url) {{
      gdsHtml = `<a href="${{esc(t.gds_url)}}" target="_blank" rel="noopener" style="color:var(--blue);text-decoration:underline;">${{esc(t.gds_text || 'Open GDS')}}</a>`;
    }} else {{
      gdsHtml = esc(t.gds_text);
    }}
    html += `<div class="detail-notes"><div class="dn-label">GDS Links</div><div class="dn-text">${{gdsHtml}}</div></div>`;
  }}
  return html;
}}
const DETAIL_H = 120;
function render() {{
  const filtered = getFiltered();
  const visWeeks = getVisibleWeeks();
  const running = filtered.filter(t => t.status === 'Running').length;
  const scheduled = filtered.filter(t => t.status === 'Scheduled').length;
  const avgLift = filtered.filter(t => t.gnarr !== null);
  const avg = avgLift.length ? (avgLift.reduce((s, t) => s + t.gnarr, 0) / avgLift.length * 100).toFixed(2) : '\\u2014';
  let statsText = `<span>${{filtered.length}} tests</span><span>${{running}} running</span>`;
  if (scheduled) statsText += `<span>${{scheduled}} scheduled</span>`;
  statsText += `<span>Avg GNARR Lift: ${{avg}}%</span><span style="opacity:0.6;">Click a row for full details</span>`;
  document.getElementById('statsBar').innerHTML = statsText;
  renderRevenueStrip(filtered);
  const lp = document.getElementById('leftPanel');
  lp.innerHTML = `<div class="left-header"><div>Product</div><div>Test Name</div><div class="col-r">QGNARR</div><div class="col-r">GNARR</div><div class="col-r">CTR</div><div class="col-r">Units</div><div>Winner</div></div>`;
  // Use global gProductMedians for compact left-panel hover tooltips
  const productMedians = gProductMedians;
  filtered.forEach((t, i) => {{
    const bg = i % 2 ? 'background:var(--row-alt);' : '';
    let winnerHtml = '\\u2014';
    if (t.winner) {{
      if (t.winner.startsWith('Challenger')) winnerHtml = `<span class="winner-challenger">${{t.winner}}</span>`;
      else if (t.winner === 'Control') winnerHtml = `<span class="winner-control">Control</span>`;
      else if (t.winner === 'Rollout') winnerHtml = `<span class="winner-rollout">Rollout</span>`;
      else winnerHtml = esc(t.winner);
    }}
    if (t.status === 'Running') winnerHtml = `<span class="status-running">Running</span>`;
    if (t.status === 'Scheduled') winnerHtml = `<span class="status-scheduled">Scheduled</span>`;
    // Build metric columns with median hover tooltip
    const pm = productMedians[t.product] || {{}};
    function mcolTip(key, val, isCurrency) {{
      const med = pm[key];
      if (val === null || med === null || med === undefined) return '';
      if (med === 0) return esc(t.product) + ' median: ' + (isCurrency ? fmtDollar(0) : '0.00%');
      const diff = ((val - med) / Math.abs(med)) * 100;
      const sign = diff > 0 ? '+' : '';
      const arrow = diff > 0 ? '\\u25B2' : diff < 0 ? '\\u25BC' : '\\u25CF';
      const medDisplay = isCurrency ? fmtDollar(med) : fmtPct(med);
      const clr = diff > 0 ? '#22c55e' : diff < 0 ? '#ef4444' : '#94a3b8';
      return `<span style="color:${{clr}}">${{arrow}} ${{sign}}${{diff.toFixed(0)}}%</span> vs ${{esc(t.product)}} median (${{medDisplay}})`;
    }}
    function mcol(key, val, isCurrency) {{
      const tipHtml = mcolTip(key, val, isCurrency);
      const tipAttr = tipHtml ? ` data-mctip="${{tipHtml.replace(/"/g, '&quot;')}}"` : '';
      const cls = val === null ? 'neutral' : val > 0 ? 'pos' : val < 0 ? 'neg' : 'neutral';
      const display = val === null ? '\\u2014' : isCurrency ? fmtDollar(val) : fmtPct(val);
      return `<div class="mcol ${{cls}}"${{tipAttr}}>${{display}}</div>`;
    }}
    const row = document.createElement('div');
    row.className = 'left-row';
    row.dataset.idx = i;
    row.setAttribute('style', bg + `height:${{ROW_H}}px;`);
    row.innerHTML = `<div><span class="cell-main"><span class="expand-icon">\\u25B6</span>${{esc(t.product)}}</span></div><div><span class="cell-main" title="${{esc(t.name)}}">${{esc(t.name)}}</span></div>${{mcol('qgnarr', t.qgnarr, true)}}${{mcol('gnarr', t.gnarr, false)}}${{mcol('ctr', t.ctr, false)}}${{mcol('units_pct', t.units_pct, false)}}<div style="align-items:center;"><span>${{winnerHtml}}</span></div>`;
    lp.appendChild(row);
    // Detail row (hidden by default)
    const detail = document.createElement('div');
    detail.className = 'detail-row';
    detail.id = 'detail-left-' + i;
    detail.innerHTML = `<div class="detail-left">${{buildDetailHtml(t)}}</div>`;
    detail.style.minHeight = '0';
    lp.appendChild(detail);
  }});
  const rp = document.getElementById('rightPanel');
  const fqGroups = {{}};
  visWeeks.forEach(w => {{ if (!fqGroups[w.fq]) fqGroups[w.fq] = []; fqGroups[w.fq].push(w); }});
  const fqColors = ['#1F4E78', '#2E75B6', '#4472C4', '#5B9BD5'];
  let fqHtml = '<div class="fq-row">', ci = 0;
  Object.entries(fqGroups).forEach(([fq, wks]) => {{
    const w = wks.length * CELL_W;
    fqHtml += `<div class="fq-label" style="width:${{w}}px;min-width:${{w}}px;background:${{fqColors[ci % 4]}};">${{fq}}</div>`;
    ci++;
  }});
  fqHtml += '</div>';
  const _now = Date.now();
  const _DAY_MS = 86400000;
  let wkHtml = '<div class="week-row">';
  visWeeks.forEach(w => {{
    const mon = w.start.toLocaleDateString('en-US', {{ month: 'short', day: 'numeric' }});
    const isCurr = _now >= w.start.getTime() && _now < w.end.getTime() + _DAY_MS;
    const wkBg = isCurr ? 'background:#2E5A8A;' : 'background:#4472C4;';
    wkHtml += `<div class="week-cell" style="width:${{CELL_W}}px;min-width:${{CELL_W}}px;${{wkBg}}"><span class="wk-num">W${{w.num}}</span><span>${{mon}}</span></div>`;
  }});
  wkHtml += '</div>';
  let barsHtml = '';
  filtered.forEach((t, i) => {{
    const ts = new Date(t.start), te = new Date(t.end);
    const bg = i % 2 ? 'background:var(--row-alt);' : '';
    let rowHtml = `<div class="bar-row" style="${{bg}}height:${{ROW_H}}px;" data-idx="${{i}}">`;
    visWeeks.forEach((w, wi) => {{
      const overlap = ts <= w.end && te >= w.start;
      const isCurr = _now >= w.start.getTime() && _now < w.end.getTime() + _DAY_MS;
      const cls = w.num === 1 ? 'bar-cell qtr-start' : 'bar-cell';
      const cellBg = isCurr ? 'background:rgba(46,90,138,0.10);' : '';
      rowHtml += `<div class="${{cls}}" style="width:${{CELL_W}}px;min-width:${{CELL_W}}px;height:${{ROW_H}}px;${{cellBg}}">`;
      if (overlap) {{
        const barStart = Math.max(ts.getTime(), w.start.getTime());
        const barEnd = Math.min(te.getTime(), w.end.getTime());
        const left = ((barStart - w.start.getTime()) / (7 * 86400000)) * CELL_W;
        const right = ((w.end.getTime() - barEnd) / (7 * 86400000)) * CELL_W;
        const fillClass = t.status === 'Scheduled' ? 'scheduled' : t.status === 'Running' ? 'running' : 'complete';
        rowHtml += `<div class="bar-fill ${{fillClass}}" style="left:${{Math.round(left)}}px;right:${{Math.round(right)}}px;"></div>`;
      }}
      rowHtml += '</div>';
    }});
    rowHtml += '</div>';
    // Matching detail spacer row for the right panel (keeps scroll in sync)
    rowHtml += `<div class="detail-row" id="detail-right-${{i}}" style="min-height:0;"></div>`;
    barsHtml += rowHtml;
  }});
  // Today indicator line for results roadmap
  let ganttTodayHtml = '';
  for (let i = 0; i < visWeeks.length; i++) {{
    const w = visWeeks[i];
    if (_now >= w.start.getTime() && _now < w.end.getTime() + _DAY_MS) {{
      const frac = (_now - w.start.getTime()) / (7 * _DAY_MS);
      const todayX = i * CELL_W + frac * CELL_W;
      ganttTodayHtml = `<div style="position:absolute;left:${{todayX}}px;top:0;bottom:0;width:2px;background:var(--red);z-index:8;pointer-events:none;">
        <div style="position:absolute;top:0;left:-20px;width:42px;text-align:center;background:var(--red);color:white;font-size:9px;font-weight:700;padding:1px 4px;border-radius:0 0 4px 4px;">Today</div>
      </div>`;
      break;
    }}
  }}
  rp.innerHTML = `<div class="timeline-header">${{fqHtml}}${{wkHtml}}</div><div style="position:relative;">${{barsHtml}}${{ganttTodayHtml}}</div>`;
  // Sync header heights so rows align
  requestAnimationFrame(() => {{
    const rpHdr = rp.querySelector('.timeline-header');
    const lpHdr = lp.querySelector('.left-header');
    if (rpHdr && lpHdr) {{
      const rpH = rpHdr.offsetHeight;
      const lpH = lpHdr.offsetHeight;
      const maxH = Math.max(rpH, lpH);
      lpHdr.style.height = maxH + 'px';
      rpHdr.style.height = maxH + 'px';
    }}
  }});
  lp.onscroll = () => {{ rp.scrollTop = lp.scrollTop; }};
  rp.onscroll = () => {{ lp.scrollTop = rp.scrollTop; }};
  // Metric column hover tooltips (fixed-positioned, avoids overflow clipping)
  const mcTipEl = document.getElementById('mcTip');
  lp.addEventListener('mouseover', (e) => {{
    const col = e.target.closest('.mcol');
    if (!col || !col.dataset.mctip) {{ mcTipEl.style.display = 'none'; return; }}
    mcTipEl.innerHTML = col.dataset.mctip;
    mcTipEl.style.display = 'block';
    const r = col.getBoundingClientRect();
    mcTipEl.style.left = (r.left + r.width / 2 - mcTipEl.offsetWidth / 2) + 'px';
    mcTipEl.style.top = (r.top - mcTipEl.offsetHeight - 8) + 'px';
  }});
  lp.addEventListener('mouseout', (e) => {{
    const col = e.target.closest('.mcol');
    if (col) mcTipEl.style.display = 'none';
  }});
  // Click to expand/collapse detail rows
  lp.querySelectorAll('.left-row').forEach((row) => {{
    row.addEventListener('click', () => {{
      const idx = row.dataset.idx;
      const dlLeft = document.getElementById('detail-left-' + idx);
      const dlRight = document.getElementById('detail-right-' + idx);
      const isOpen = dlLeft.classList.contains('open');
      // Close all other open details
      document.querySelectorAll('.detail-row.open').forEach(d => {{
        d.classList.remove('open');
        d.style.height = '0';
      }});
      document.querySelectorAll('.left-row.expanded').forEach(r => r.classList.remove('expanded'));
      if (!isOpen) {{
        dlLeft.classList.add('open');
        dlRight.classList.add('open');
        const h = dlLeft.querySelector('.detail-left').scrollHeight + 20;
        dlLeft.style.height = h + 'px';
        dlRight.style.height = h + 'px';
        row.classList.add('expanded');
      }}
    }});
  }});
  // Also allow clicking bar rows to expand
  rp.querySelectorAll('.bar-row').forEach((row) => {{
    row.addEventListener('click', () => {{
      const idx = row.dataset.idx;
      const leftRow = lp.querySelector(`.left-row[data-idx="${{idx}}"]`);
      if (leftRow) leftRow.click();
    }});
  }});
  // Hover sync between panels
  rp.querySelectorAll('.bar-row').forEach((row, i) => {{
    const t = filtered[i];
    row.addEventListener('mouseenter', (e) => {{
      tooltip.style.display = 'block';
      tooltip.innerHTML = buildTooltipHtml(t);
      const lr = lp.querySelectorAll('.left-row')[i]; if (lr) lr.style.background = 'var(--blue-light)';
    }});
    row.addEventListener('mousemove', (e) => {{
      tooltip.style.left = Math.min(e.clientX + 12, window.innerWidth - 400) + 'px';
      tooltip.style.top = Math.min(e.clientY + 12, window.innerHeight - 250) + 'px';
    }});
    row.addEventListener('mouseleave', () => {{
      tooltip.style.display = 'none';
      const lr = lp.querySelectorAll('.left-row')[i]; if (lr) lr.style.background = '';
    }});
  }});
}}
function renderActive() {{
  const active = document.querySelector('.view-tab.active');
  const view = active ? active.dataset.view : 'gantt';
  render();
  if (view === 'scatter') renderScatter();
  if (view === 'list') renderList();
  if (view === 'plan') {{ planLoadSaved().then(() => renderPlan()); }}
}}
prodSel.addEventListener('change', () => {{ activeSurface = prodSel.value; renderActive(); }});
document.getElementById('statusFilter').addEventListener('change', renderActive);
document.getElementById('searchInput').addEventListener('input', renderActive);
document.getElementById('scatterMetric').addEventListener('change', renderScatter);

// Surface toggle
document.querySelectorAll('.surface-btn').forEach(btn => {{
  btn.addEventListener('click', () => {{
    document.querySelectorAll('.surface-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    activeSurface = btn.dataset.val;
    renderActive();
  }});
}});
document.getElementById('zoomSlider').addEventListener('input', (e) => {{ CELL_W = parseInt(e.target.value); render(); }});
// First render to establish DOM layout, then measure right panel and fit 13 weeks
render();
requestAnimationFrame(() => {{
  const rp = document.getElementById('rightPanel');
  const rpW = rp.clientWidth;
  if (rpW > 100) {{
    CELL_W = Math.max(8, Math.min(80, Math.floor(rpW / 13)));
    document.getElementById('zoomSlider').value = CELL_W;
    render();
  }}
}});
// ── View tab switching ──
document.querySelectorAll('.view-tab').forEach(tab => {{
  tab.addEventListener('click', () => {{
    document.querySelectorAll('.view-tab').forEach(t => t.classList.remove('active'));
    document.querySelectorAll('.view-panel').forEach(p => p.classList.remove('active'));
    tab.classList.add('active');
    document.getElementById(tab.dataset.view + 'View').classList.add('active');
    if (tab.dataset.view === 'scatter') renderScatter();
    if (tab.dataset.view === 'list') renderList();
    if (tab.dataset.view === 'plan') {{ planLoadSaved().then(() => renderPlan()); }}
    if (tab.dataset.view === 'velocity') renderVelocity();
  }});
}});
// ── Experiment List Table ──
let listSortCol = 'start';
let listSortDir = -1; // -1 = descending (newest first)
function renderList() {{
  const container = document.getElementById('listContainer');
  // Apply same filters as Gantt (product, status, search, quarter)
  const filtered = getFiltered();
  // Sort
  const sorted = filtered.slice().sort((a, b) => {{
    let va = a[listSortCol], vb = b[listSortCol];
    if (listSortCol === 'start' || listSortCol === 'end') {{
      va = va ? new Date(va).getTime() : 0;
      vb = vb ? new Date(vb).getTime() : 0;
    }} else if (listSortCol === 'qgnarr' || listSortCol === 'gnarr' || listSortCol === 'ctr' || listSortCol === 'units_pct') {{
      va = va !== null ? va : -Infinity;
      vb = vb !== null ? vb : -Infinity;
    }} else {{
      va = va ? String(va).toLowerCase() : '';
      vb = vb ? String(vb).toLowerCase() : '';
    }}
    if (va < vb) return -1 * listSortDir;
    if (va > vb) return 1 * listSortDir;
    return 0;
  }});
  const cols = [
    {{ key: 'product', label: 'Product', cls: '' }},
    {{ key: 'name', label: 'Test Name', cls: 'col-name' }},
    {{ key: 'fq', label: 'Quarter', cls: '' }},
    {{ key: 'start', label: 'Start', cls: '' }},
    {{ key: 'end', label: 'End', cls: '' }},
    {{ key: 'status', label: 'Status', cls: '' }},
    {{ key: 'winner', label: 'Winner', cls: '' }},
    {{ key: 'qgnarr', label: 'QGNARR', cls: 'col-num' }},
    {{ key: 'gnarr', label: 'GNARR Lift', cls: 'col-num' }},
    {{ key: 'ctr', label: 'CTR', cls: 'col-num' }},
    {{ key: 'units_pct', label: 'Units %', cls: 'col-num' }},
    {{ key: 'gds', label: 'GDS', cls: '' }},
    {{ key: 'details', label: 'Details', cls: 'col-detail' }},
    {{ key: 'pm_commentary', label: 'PM Commentary', cls: 'col-detail' }}
  ];
  let html = '<table class="exp-table"><thead><tr>';
  cols.forEach(c => {{
    const isSorted = listSortCol === c.key;
    const arrow = isSorted ? (listSortDir === 1 ? '\\u25B2' : '\\u25BC') : '\\u25B2';
    html += `<th class="${{isSorted ? 'sorted' : ''}}" data-col="${{c.key}}">${{c.label}}<span class="sort-arrow">${{arrow}}</span></th>`;
  }});
  html += '</tr></thead><tbody>';
  if (sorted.length === 0) {{
    html += `<tr><td colspan="${{cols.length}}" style="text-align:center;padding:24px;color:var(--text-muted);">No experiments match the current filters</td></tr>`;
  }}
  sorted.forEach(t => {{
    const winCls = t.winner && t.winner.startsWith('Challenger') ? 'winner-challenger' : t.winner === 'Control' ? 'winner-control' : t.winner === 'Rollout' ? 'winner-rollout' : (t.status === 'Running' ? 'status-running' : (t.status === 'Scheduled' ? 'status-scheduled' : ''));
    const winHtml = t.winner ? `<span class="${{winCls}}">${{t.winner}}</span>` : (t.status === 'Running' ? '<span class="status-running">Running</span>' : (t.status === 'Scheduled' ? '<span class="status-scheduled">Scheduled</span>' : '\\u2014'));
    const statusHtml = t.status === 'Running' ? '<span class="status-running">Running</span>' : (t.status === 'Scheduled' ? '<span class="status-scheduled">Scheduled</span>' : t.status);
    const gnarrColor = t.gnarr > 0 ? 'var(--green)' : t.gnarr < 0 ? 'var(--red)' : 'inherit';
    const gdsHtml = t.gds_url ? `<a href="${{t.gds_url}}" target="_blank">${{esc(t.gds_text) || 'Link'}}</a>` : (t.gds_text ? esc(t.gds_text) : '\\u2014');
    const detailHtml = t.details ? `<div class="col-detail-text">${{esc(t.details)}}</div>` : '\\u2014';
    const pmHtml = t.pm_commentary ? `<div class="col-detail-text">${{esc(t.pm_commentary)}}</div>` : '\\u2014';
    html += `<tr>
      <td>${{esc(t.product)}}</td>
      <td class="col-name">${{esc(t.name)}}</td>
      <td>${{t.fq}}</td>
      <td>${{fmtDate(t.start)}}</td>
      <td>${{(t.status === 'Running' || t.status === 'Scheduled') ? '\\u2014' : fmtDate(t.end)}}</td>
      <td>${{statusHtml}}</td>
      <td style="white-space:nowrap;">${{winHtml}}</td>
      <td class="col-num">${{fmtDollar(t.qgnarr)}}</td>
      <td class="col-num" style="color:${{gnarrColor}}">${{fmtPct(t.gnarr)}}</td>
      <td class="col-num">${{fmtPct(t.ctr)}}</td>
      <td class="col-num">${{fmtPct(t.units_pct)}}</td>
      <td>${{gdsHtml}}</td>
      <td class="col-detail">${{detailHtml}}</td>
      <td class="col-detail">${{pmHtml}}</td>
    </tr>`;
  }});
  html += '</tbody></table>';
  container.innerHTML = html;
  // Wire up sortable headers
  container.querySelectorAll('.exp-table th').forEach(th => {{
    th.addEventListener('click', () => {{
      const col = th.dataset.col;
      if (listSortCol === col) {{
        listSortDir *= -1;
      }} else {{
        listSortCol = col;
        listSortDir = (col === 'start' || col === 'end' || col === 'qgnarr' || col === 'gnarr' || col === 'ctr' || col === 'units_pct') ? -1 : 1;
      }}
      renderList();
    }});
  }});
  // Click detail text to expand/collapse
  container.querySelectorAll('.col-detail-text').forEach(el => {{
    el.style.cursor = 'pointer';
    el.addEventListener('click', () => el.classList.toggle('expanded'));
  }});
}}
// ── Scatter Chart ──
function renderScatter() {{
  const svg = document.getElementById('scatterSvg');
  const container = document.getElementById('scatterContainer');
  const W = container.clientWidth;
  const H = container.clientHeight;
  if (W < 100 || H < 100) {{
    requestAnimationFrame(renderScatter);
    return;
  }}
  const scatterMetric = document.getElementById('scatterMetric').value;
  const metricLabels = {{ gnarr: 'GNARR Lift', ctr: 'CTR', units_pct: 'Units' }};
  const metricLabel = metricLabels[scatterMetric] || 'Lift';
  svg.setAttribute('viewBox', `0 0 ${{W}} ${{H}}`);
  svg.setAttribute('preserveAspectRatio', 'xMidYMid meet');
  const pad = {{ top: 40, right: 40, bottom: 60, left: 70 }};
  const plotW = W - pad.left - pad.right;
  const plotH = H - pad.top - pad.bottom;
  // Rolling 24-month window — ignore quarter filter, use product/status/search only
  const now = new Date();
  const windowStart = new Date(now.getFullYear() - 2, now.getMonth(), 1);
  const windowEnd = new Date(now.getFullYear(), now.getMonth() + 1, 0);
  // Apply product/status/search filters but NOT quarter filter
  const product = activeSurface;
  const status = document.getElementById('statusFilter').value;
  const search = document.getElementById('searchInput').value.toLowerCase();
  const filtered = TESTS.filter(t => {{
    if (t[scatterMetric] === null || t[scatterMetric] === undefined) return false;
    if (t.status !== 'Complete') return false;
    const dateVal = t.end || t.start;
    const te = dateVal ? new Date(dateVal) : null;
    if (!te || isNaN(te.getTime())) return false;
    if (te < windowStart || te > windowEnd) return false;
    if (product !== 'All' && t.product !== product) return false;
    if (status !== 'All' && t.status !== status) return false;
    if (search && !t.name.toLowerCase().includes(search)) return false;
    return true;
  }});
  if (filtered.length === 0) {{ const surfaceLabel = product === 'All' ? 'any surface' : product;
    svg.innerHTML = `<text x="50%" y="50%" text-anchor="middle" fill="#718096" font-size="14">No completed tests with ${{metricLabel}} data for ${{surfaceLabel}} in the last 24 months</text>`; return; }}
  // Fixed X-axis: windowStart to windowEnd (24 months)
  const minDate = windowStart.getTime();
  const maxDate = windowEnd.getTime();
  const dateRange = maxDate - minDate;
  // Y-axis from data
  const lifts = filtered.map(t => t[scatterMetric] * 100);
  const maxLift = Math.ceil(Math.max(...lifts, 1));
  const minLift = Math.floor(Math.min(...lifts, -1));
  const liftRange = maxLift - minLift || 1;
  const avg = lifts.reduce((s, v) => s + v, 0) / lifts.length;
  function xPos(d) {{ return pad.left + ((d - minDate) / dateRange) * plotW; }}
  function yPos(v) {{ return pad.top + ((maxLift - v) / liftRange) * plotH; }}
  function dotColor(lift) {{
    if (lift >= 0.5) return '#006100';
    if (lift <= -0.5) return '#9C0006';
    return '#BF8F00';
  }}
  function dotFill(lift) {{
    if (lift >= 0.5) return '#22c55e';
    if (lift <= -0.5) return '#ef4444';
    return '#FFC000';
  }}
  let html = '';
  // Y-axis grid lines
  for (let v = Math.ceil(minLift); v <= Math.floor(maxLift); v += 1) {{
    if (v === 0) continue;
    const y = yPos(v);
    html += `<line x1="${{pad.left}}" y1="${{y}}" x2="${{pad.left + plotW}}" y2="${{y}}" class="scatter-grid"/>`;
    html += `<text x="${{pad.left - 8}}" y="${{y + 3}}" text-anchor="end" class="scatter-axis">${{v}}%</text>`;
  }}
  // Zero line
  if (minLift <= 0 && maxLift >= 0) {{
    html += `<line x1="${{pad.left}}" y1="${{yPos(0)}}" x2="${{pad.left + plotW}}" y2="${{yPos(0)}}" class="scatter-zero"/>`;
    html += `<text x="${{pad.left - 8}}" y="${{yPos(0) + 3}}" text-anchor="end" class="scatter-axis" style="font-weight:600;">0%</text>`;
  }}
  // X-axis: monthly ticks across the full 24-month window
  const tickDate = new Date(windowStart);
  while (tickDate <= windowEnd) {{
    const x = xPos(tickDate.getTime());
    if (x >= pad.left && x <= pad.left + plotW) {{
      const label = tickDate.toLocaleDateString('en-US', {{ month: 'short', year: '2-digit' }});
      html += `<line x1="${{x}}" y1="${{pad.top + plotH}}" x2="${{x}}" y2="${{pad.top + plotH + 6}}" stroke="#a0aec0"/>`;
      html += `<text x="${{x}}" y="${{pad.top + plotH + 20}}" text-anchor="middle" class="scatter-axis">${{label}}</text>`;
    }}
    tickDate.setMonth(tickDate.getMonth() + 1);
  }}
  // Axis labels
  html += `<text x="${{pad.left + plotW / 2}}" y="${{H - 5}}" text-anchor="middle" class="scatter-axis-label">Test End Date</text>`;
  html += `<text x="15" y="${{pad.top + plotH / 2}}" text-anchor="middle" class="scatter-axis-label" transform="rotate(-90,15,${{pad.top + plotH / 2}})">${{metricLabel}} %</text>`;
  // Avg dashed line
  const avgY = yPos(avg);
  html += `<line x1="${{pad.left}}" y1="${{avgY}}" x2="${{pad.left + plotW}}" y2="${{avgY}}" stroke="#2E75B6" stroke-width="1.5" class="scatter-avg-line"/>`;
  html += `<text x="${{pad.left + 6}}" y="${{avgY - 6}}" class="scatter-avg-label">Overall Avg: ${{avg.toFixed(1)}}%</text>`;
  // Plot border
  html += `<rect x="${{pad.left}}" y="${{pad.top}}" width="${{plotW}}" height="${{plotH}}" fill="none" stroke="#e2e8f0"/>`;
  // Dots
  filtered.forEach((t, i) => {{
    const endDate = new Date(t.end).getTime();
    const lift = t[scatterMetric] * 100;
    const x = xPos(endDate);
    const y = yPos(lift);
    const fill = dotFill(lift);
    const liftStr = lift >= 0 ? '+' + lift.toFixed(1) + '%' : lift.toFixed(1) + '%';
    // Jitter overlapping dots slightly
    const jitter = (i % 3 - 1) * 3;
    html += `<circle class="scatter-dot" cx="${{x + jitter}}" cy="${{y}}" r="5" fill="${{fill}}" stroke="white" stroke-width="1.5" data-idx="${{i}}"/>`;
    html += `<text class="scatter-label" x="${{x + jitter}}" y="${{y - 9}}" text-anchor="middle" fill="${{dotColor(lift)}}">${{liftStr}}</text>`;
  }});
  // Legend
  const lx = pad.left + plotW - 340;
  const ly = pad.top + 10;
  html += `<rect x="${{lx}}" y="${{ly}}" width="340" height="26" rx="4" fill="white" fill-opacity="0.9" stroke="#e2e8f0"/>`;
  html += `<circle cx="${{lx + 14}}" cy="${{ly + 13}}" r="5" fill="#22c55e"/>`;
  html += `<text x="${{lx + 24}}" y="${{ly + 17}}" font-size="10" fill="#1a202c">Positive (\\u22650.5%)</text>`;
  html += `<circle cx="${{lx + 134}}" cy="${{ly + 13}}" r="5" fill="#ef4444"/>`;
  html += `<text x="${{lx + 144}}" y="${{ly + 17}}" font-size="10" fill="#1a202c">Negative (\\u2264-0.5%)</text>`;
  html += `<circle cx="${{lx + 264}}" cy="${{ly + 13}}" r="5" fill="#FFC000"/>`;
  html += `<text x="${{lx + 274}}" y="${{ly + 17}}" font-size="10" fill="#1a202c">Flat</text>`;
  svg.innerHTML = html;
  // Hover tooltips on dots
  svg.querySelectorAll('.scatter-dot').forEach((dot, i) => {{
    const t = filtered[i];
    dot.addEventListener('mouseenter', (e) => {{
      tooltip.style.display = 'block';
      tooltip.innerHTML = buildTooltipHtml(t);
    }});
    dot.addEventListener('mousemove', (e) => {{
      tooltip.style.left = Math.min(e.clientX + 12, window.innerWidth - 400) + 'px';
      tooltip.style.top = Math.min(e.clientY + 12, window.innerHeight - 250) + 'px';
    }});
    dot.addEventListener('mouseleave', () => {{ tooltip.style.display = 'none'; }});
  }});
}}
// ── Velocity Chart ──
let _velChart = null;
function renderVelocity() {{
  const canvas = document.getElementById('velocityCanvas');
  const container = document.getElementById('velocityContainer');
  if (!container.clientWidth || !container.clientHeight) {{
    requestAnimationFrame(renderVelocity);
    return;
  }}
  // Bin tests by start month (18-month window ending current month)
  const now = new Date();
  const windowStart = new Date(now.getFullYear() - 1, now.getMonth() - 5, 1); // ~18 months back
  const windowEnd = new Date(now.getFullYear(), now.getMonth() + 1, 0);
  // Build month labels and counts
  const months = [];
  const dt = new Date(windowStart);
  while (dt <= windowEnd) {{
    months.push({{ label: dt.toLocaleString('en', {{ month: 'short' }}) + ' ' + String(dt.getFullYear()).slice(2), year: dt.getFullYear(), month: dt.getMonth(), count: 0 }});
    dt.setMonth(dt.getMonth() + 1);
  }}
  // Dedup: collapse language variants (same name minus trailing lang code)
  const seen = new Set();
  TESTS.forEach(t => {{
    const s = new Date(t.start);
    if (s < windowStart || s > windowEnd) return;
    // Simple dedup: strip trailing " - XX" language codes
    const baseName = t.name.replace(/\\s*[-–]\\s*[A-Z]{{2}}$/i, '').trim();
    const key = baseName + '|' + s.getFullYear() + '-' + s.getMonth();
    if (seen.has(key)) return;
    seen.add(key);
    const mi = months.findIndex(m => m.year === s.getFullYear() && m.month === s.getMonth());
    if (mi >= 0) months[mi].count++;
  }});
  const labels = months.map(m => m.label);
  const counts = months.map(m => m.count);
  // 3-month rolling average
  const rolling = counts.map((v, i) => {{
    if (i < 2) return null;
    return Math.round(((counts[i - 2] + counts[i - 1] + v) / 3) * 10) / 10;
  }});
  // Multi-tcat phases (by month index)
  function monthIdx(y, m) {{ return months.findIndex(mo => mo.year === y && mo.month === m); }}
  const phases = [
    {{ name: 'Multi-tcat testing', start: monthIdx(2025, 5), end: monthIdx(2025, 9), color: 'rgba(59,130,246,0.10)', border: 'rgba(59,130,246,0.3)' }},
    {{ name: 'Multi-tcat 100% rollout', start: monthIdx(2025, 9), end: monthIdx(2025, 11), color: 'rgba(34,197,94,0.10)', border: 'rgba(34,197,94,0.3)' }},
    {{ name: 'Validation (reduced)', start: monthIdx(2026, 0), end: monthIdx(2026, 1), color: 'rgba(234,179,8,0.10)', border: 'rgba(234,179,8,0.3)' }},
    {{ name: 'Post-shutdown', start: monthIdx(2026, 2), end: months.length - 1, color: 'rgba(239,68,68,0.08)', border: 'rgba(239,68,68,0.25)' }}
  ];
  // Single-tcat baseline: avg of Jan-May 2025
  const baselineMonths = months.filter(m => m.year === 2025 && m.month >= 0 && m.month <= 4);
  const baseline = baselineMonths.length ? Math.round((baselineMonths.reduce((s, m) => s + m.count, 0) / baselineMonths.length) * 10) / 10 : 1.8;
  // Phase annotation boxes
  const phaseBoxes = phases.map(p => ({{
    type: 'box', xMin: p.start - 0.5, xMax: p.end + 0.5,
    backgroundColor: p.color, borderColor: p.border, borderWidth: 1,
    label: {{ display: true, content: p.name, position: 'start', font: {{ size: 10, weight: '600' }}, color: '#475569', padding: 4 }}
  }}));
  // Baseline annotation
  phaseBoxes.push({{
    type: 'line', yMin: baseline, yMax: baseline, borderColor: '#94a3b8', borderWidth: 2,
    borderDash: [6, 4],
    label: {{ display: true, content: 'Single-tcat baseline (' + baseline + '/mo)', position: 'end', font: {{ size: 10 }}, color: '#64748b', backgroundColor: 'rgba(255,255,255,0.85)', padding: 4 }}
  }});
  if (_velChart) _velChart.destroy();
  _velChart = new Chart(canvas, {{
    type: 'bar',
    data: {{
      labels: labels,
      datasets: [
        {{
          label: 'Experiments launched',
          data: counts,
          backgroundColor: counts.map((v, i) => {{
            const m = months[i];
            const phase = phases.find(p => i >= p.start && i <= p.end);
            if (phase) {{
              if (phase.name.includes('testing')) return 'rgba(59,130,246,0.7)';
              if (phase.name.includes('rollout')) return 'rgba(34,197,94,0.7)';
              if (phase.name.includes('Validation')) return 'rgba(234,179,8,0.7)';
              if (phase.name.includes('shutdown')) return 'rgba(239,68,68,0.5)';
            }}
            return 'rgba(100,116,139,0.5)';
          }}),
          borderColor: counts.map((v, i) => {{
            const phase = phases.find(p => i >= p.start && i <= p.end);
            if (phase) {{
              if (phase.name.includes('testing')) return '#3b82f6';
              if (phase.name.includes('rollout')) return '#22c55e';
              if (phase.name.includes('Validation')) return '#eab308';
              if (phase.name.includes('shutdown')) return '#ef4444';
            }}
            return '#64748b';
          }}),
          borderWidth: 1, borderRadius: 3, order: 2
        }},
        {{
          label: '3-mo rolling avg',
          data: rolling,
          type: 'line',
          borderColor: '#1e40af',
          backgroundColor: 'transparent',
          borderWidth: 2,
          pointRadius: 2,
          pointBackgroundColor: '#1e40af',
          tension: 0.3,
          spanGaps: false,
          order: 1
        }}
      ]
    }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      plugins: {{
        legend: {{ display: true, position: 'top', labels: {{ font: {{ size: 11 }}, usePointStyle: true, padding: 16 }} }},
        annotation: {{ annotations: phaseBoxes }},
        tooltip: {{
          callbacks: {{
            afterLabel: function(ctx) {{
              if (ctx.datasetIndex === 0) {{
                const i = ctx.dataIndex;
                const phase = phases.find(p => i >= p.start && i <= p.end);
                const phaseName = phase ? phase.name : 'Single-tcat infra';
                const mult = baseline > 0 ? (counts[i] / baseline).toFixed(1) : '—';
                return phaseName + ' (' + mult + '× baseline)';
              }}
              return '';
            }}
          }}
        }}
      }},
      scales: {{
        x: {{ grid: {{ display: false }}, ticks: {{ font: {{ size: 10 }}, maxRotation: 45 }} }},
        y: {{ beginAtZero: true, ticks: {{ stepSize: 1, font: {{ size: 11 }} }}, title: {{ display: true, text: 'Experiments launched', font: {{ size: 12, weight: '600' }} }} }}
      }}
    }}
  }});
}}
// ── Planning Roadmap ──
const BACKLOG = RAW_DATA.backlog || [];
const AVG_DURATION = RAW_DATA.avgDuration || {{}};
const PLAN_CELL = 50; // px per week
const PLAN_ROW_H = 40;
let plannedItems = []; // {{ id, name, product, startOffset, duration, startDate }}
let _planLoaded = false;
let _planDirty = false;
const GH_REPO = 'matthewgshep/infra-testing-roadmap';
const GH_FILE = 'plan.json';
const GH_BRANCH = 'gh-pages';  // plan.json lives on the published branch

function ghGetToken() {{ return localStorage.getItem('gh_plan_token') || ''; }}
function ghSetToken(t) {{ localStorage.setItem('gh_plan_token', t); }}
function ghClearToken() {{ localStorage.removeItem('gh_plan_token'); }}

function ghPromptToken() {{
  const t = prompt('Enter a GitHub Personal Access Token with Contents read/write permission for ' + GH_REPO + ':\\n\\n(Fine-grained token → Repository access → Only select ' + GH_REPO.split('/')[1] + ' → Contents: Read and write)');
  if (t && t.trim()) {{ ghSetToken(t.trim()); return true; }}
  return false;
}}

function planMarkDirty() {{
  _planDirty = true;
  const btn = document.getElementById('planSaveBtn');
  const statusEl = document.getElementById('planSaveStatus');
  if (btn) {{ btn.disabled = false; btn.style.opacity = '1'; }}
  if (statusEl) {{ statusEl.textContent = 'Unsaved changes'; statusEl.style.color = 'var(--gold)'; }}
}}

function planBuildPayload() {{
  const quarters = planGetQuarters();
  const weeks = planWeeks(quarters);
  const timelineStart = weeks[0].start.getTime();
  return plannedItems.map(p => {{
    const startMs = timelineStart + p.startOffset * 7 * 86400000;
    const endMs = startMs + p.duration * 7 * 86400000;
    const sd = new Date(startMs);
    const ed = new Date(endMs);
    // Find which FQ week the start falls in
    const wi = Math.floor(p.startOffset);
    const fqWeek = weeks[wi] ? weeks[wi].fq + ' W' + weeks[wi].num : '';
    return {{
      name: p.name,
      product: p.product,
      startDate: sd.toISOString().slice(0, 10),
      endDate: ed.toISOString().slice(0, 10),
      durationWeeks: Math.round(p.duration * 10) / 10,
      fqWeek: fqWeek
    }};
  }});
}}

function planDateToOffset(dateStr, weeks) {{
  // Returns fractional week offset from timeline start, or -1 if outside range
  const d = new Date(dateStr).getTime();
  const timelineStart = weeks[0].start.getTime();
  const timelineEnd = weeks[weeks.length - 1].end.getTime() + 86400000;
  if (d < timelineStart || d > timelineEnd) return -1;
  return (d - timelineStart) / (7 * 86400000);
}}

async function planLoadSaved() {{
  if (_planLoaded) return;
  const statusEl = document.getElementById('planSaveStatus');
  try {{
    const resp = await fetch(GH_FILE + '?_=' + Date.now());
    if (!resp.ok) {{ _planLoaded = true; return; }}
    const saved = await resp.json();
    if (!Array.isArray(saved) || saved.length === 0) {{ _planLoaded = true; return; }}
    const quarters = planGetQuarters();
    const weeks = planWeeks(quarters);
    plannedItems = [];
    saved.forEach(s => {{
      const off = planDateToOffset(s.startDate, weeks);
      if (off < 0) return; // skip items outside visible range
      // Calculate duration from saved dates if both present, else use saved durationWeeks
      let dur = s.durationWeeks || 4;
      if (s.endDate && s.startDate) {{
        const daysDur = (new Date(s.endDate) - new Date(s.startDate)) / 86400000;
        if (daysDur > 0) dur = daysDur / 7;
      }}
      plannedItems.push({{
        id: Date.now() + Math.random(),
        name: s.name,
        product: s.product,
        startOffset: off,
        duration: dur,
        startDate: s.startDate
      }});
    }});
    _planLoaded = true;
    _planDirty = false;
    if (statusEl && plannedItems.length > 0) {{
      statusEl.textContent = '\u2713 Loaded ' + plannedItems.length + ' items';
      statusEl.style.color = 'var(--green)';
      setTimeout(() => {{ if (statusEl) {{ statusEl.textContent = ''; }} }}, 3000);
    }}
    renderPlan();
  }} catch(e) {{
    _planLoaded = true;
    if (statusEl) {{
      statusEl.textContent = 'Could not load saved plan';
      statusEl.style.color = 'var(--text-muted)';
    }}
  }}
}}

async function planSaveToGitHub() {{
  let token = ghGetToken();
  if (!token) {{
    if (!ghPromptToken()) return;
    token = ghGetToken();
  }}
  const statusEl = document.getElementById('planSaveStatus');
  const btn = document.getElementById('planSaveBtn');
  if (statusEl) {{ statusEl.textContent = 'Saving to GitHub...'; statusEl.style.color = 'var(--blue)'; }}
  if (btn) {{ btn.disabled = true; btn.style.opacity = '0.5'; }}

  const plan = planBuildPayload();
  const contentB64 = btoa(unescape(encodeURIComponent(JSON.stringify(plan, null, 2))));

  try {{
    // Get current SHA if file exists
    const getResp = await fetch('https://api.github.com/repos/' + GH_REPO + '/contents/' + GH_FILE + '?ref=' + GH_BRANCH, {{
      headers: {{ 'Authorization': 'Bearer ' + token, 'Accept': 'application/vnd.github.v3+json' }}
    }});
    let sha = null;
    if (getResp.ok) {{
      const gd = await getResp.json();
      sha = gd.sha;
    }}

    const body = {{
      message: 'Update experiment plan (' + plan.length + ' items)',
      content: contentB64,
      branch: GH_BRANCH
    }};
    if (sha) body.sha = sha;

    const putResp = await fetch('https://api.github.com/repos/' + GH_REPO + '/contents/' + GH_FILE, {{
      method: 'PUT',
      headers: {{
        'Authorization': 'Bearer ' + token,
        'Accept': 'application/vnd.github.v3+json',
        'Content-Type': 'application/json'
      }},
      body: JSON.stringify(body)
    }});

    if (putResp.ok) {{
      _planDirty = false;
      if (statusEl) {{ statusEl.textContent = '\u2713 Saved ' + plan.length + ' items to GitHub'; statusEl.style.color = 'var(--green)'; }}
      if (btn) {{ btn.disabled = true; btn.style.opacity = '0.5'; }}
      setTimeout(() => {{ if (statusEl) {{ statusEl.textContent = ''; }} }}, 4000);
    }} else {{
      const err = await putResp.json();
      if (putResp.status === 401 || putResp.status === 403) {{
        ghClearToken();
        if (statusEl) {{ statusEl.textContent = '\u2717 Token invalid — click Save to re-enter'; statusEl.style.color = 'var(--red)'; }}
      }} else {{
        if (statusEl) {{ statusEl.textContent = '\u2717 Save failed: ' + (err.message || putResp.status); statusEl.style.color = 'var(--red)'; }}
      }}
      if (btn) {{ btn.disabled = false; btn.style.opacity = '1'; }}
    }}
  }} catch(e) {{
    if (statusEl) {{ statusEl.textContent = '\u2717 Network error'; statusEl.style.color = 'var(--red)'; }}
    if (btn) {{ btn.disabled = false; btn.style.opacity = '1'; }}
  }}
}}

function planGetQuarters() {{
  // Current quarter + 1 future quarter
  const today = new Date();
  return ALL_QUARTERS.filter(q => new Date(q.end) >= today).slice(0, 2);
}}

function planWeeks(quarters) {{
  const weeks = [];
  quarters.forEach(q => {{
    const qs = new Date(q.start);
    for (let w = 0; w < 13; w++) {{
      const ws = new Date(qs.getTime() + w * 7 * 86400000);
      const we = new Date(ws.getTime() + 6 * 86400000);
      weeks.push({{ start: ws, end: we, num: w + 1, fq: q.label, idx: weeks.length }});
    }}
  }});
  return weeks;
}}

function renderPlan() {{
  const timeline = document.getElementById('planTimeline');
  const backlogList = document.getElementById('planBacklogList');
  const backlogCount = document.getElementById('planBacklogCount');
  const quarters = planGetQuarters();
  const weeks = planWeeks(quarters);
  if (weeks.length === 0) return;
  const totalW = weeks.length * PLAN_CELL;
  const labelW = 180;

  // Product filter
  const product = prodSel.value;

  // Find existing tests that overlap this time range, filtered by product
  const rangeStart = weeks[0].start, rangeEnd = weeks[weeks.length - 1].end;
  const existing = TESTS.filter(t => {{
    if (product !== 'All' && t.product !== product) return false;
    const ts = new Date(t.start), te = new Date(t.end);
    return ts <= rangeEnd && te >= rangeStart;
  }});

  // Build timeline header (quarter + week labels)
  let hdrHtml = `<div style="position:sticky;top:0;z-index:10;">`;
  // Quarter row — each quarter spans exactly 13 week cells
  hdrHtml += `<div style="display:flex;margin-left:${{labelW}}px;">`;
  quarters.forEach((q, qi) => {{
    // Sum the widths of the 13 week cells for this quarter (each PLAN_CELL + 1px border)
    const qWidth = 13 * PLAN_CELL;
    const borderStyle = qi < quarters.length - 1 ? 'border-right:1px solid rgba(255,255,255,0.3);' : '';
    hdrHtml += `<div style="width:${{qWidth}}px;min-width:${{qWidth}}px;box-sizing:border-box;background:var(--blue-dark);color:white;text-align:center;font-size:11px;font-weight:700;padding:4px 0;${{borderStyle}}">${{q.label}}</div>`;
  }});
  hdrHtml += `</div>`;
  // Week row
  const _planNow = Date.now();
  const _PDAY = 86400000;
  hdrHtml += `<div style="display:flex;margin-left:${{labelW}}px;">`;
  weeks.forEach(w => {{
    const mon = w.start.toLocaleDateString('en-US', {{ month: 'short' }});
    const isCurr = _planNow >= w.start.getTime() && _planNow < w.end.getTime() + _PDAY;
    const wkBg = isCurr ? '#2E5A8A' : '#4472C4';
    hdrHtml += `<div style="width:${{PLAN_CELL}}px;min-width:${{PLAN_CELL}}px;box-sizing:border-box;background:${{wkBg}};color:white;text-align:center;font-size:8px;padding:2px 0;border-right:1px solid rgba(255,255,255,0.15);"><span style="font-weight:700;font-size:9px;">W${{w.num}}</span><br>${{mon}}</div>`;
  }});
  hdrHtml += `</div></div>`;

  // Build lanes — existing tests first, then planned
  let lanesHtml = '';
  // Existing / running tests
  existing.forEach(t => {{
    const ts = new Date(t.start), te = new Date(t.end);
    const barClass = t.status === 'Scheduled' ? 'scheduled-plan' : t.status === 'Running' ? 'running' : 'existing';
    const barLeft = calcBarLeft(ts, weeks, labelW);
    const barWidth = calcBarWidth(ts, te, weeks);
    lanesHtml += `<div class="plan-lane" style="height:${{PLAN_ROW_H}}px;min-width:${{labelW + totalW}}px;">
      <div class="plan-lane-label" style="width:${{labelW}}px;font-size:10px;color:var(--text-muted);">${{esc(t.name.substring(0, 30))}}</div>
      <div class="plan-bar ${{barClass}}" style="left:${{barLeft}}px;width:${{barWidth}}px;" title="${{esc(t.name)}} (${{t.status}})">${{esc(t.name.substring(0, 20))}}</div>
    </div>`;
  }});
  // Planned tests
  plannedItems.forEach((p, pi) => {{
    const barLeft = labelW + p.startOffset * PLAN_CELL;
    const barWidth = p.duration * PLAN_CELL;
    lanesHtml += `<div class="plan-lane" style="height:${{PLAN_ROW_H}}px;min-width:${{labelW + totalW}}px;" data-plan-idx="${{pi}}">
      <div class="plan-lane-label" style="width:${{labelW}}px;">${{esc(p.name.substring(0, 28))}}</div>
      <div class="plan-bar planned" style="left:${{barLeft}}px;width:${{barWidth}}px;" data-plan-idx="${{pi}}">
        <div class="plan-resize left" data-plan-idx="${{pi}}" data-edge="left"></div>
        ${{esc(p.name.substring(0, 20))}}
        <div class="plan-resize right" data-plan-idx="${{pi}}" data-edge="right"></div>
        <div class="plan-remove" data-plan-idx="${{pi}}">\\u00D7</div>
      </div>
    </div>`;
  }});
  // Empty drop zone at bottom
  lanesHtml += `<div class="plan-lane" style="height:${{PLAN_ROW_H * 3}}px;min-width:${{labelW + totalW}}px;border-bottom:none;">
    <div class="plan-lane-label" style="width:${{labelW}}px;color:var(--text-light);font-style:italic;font-size:10px;">Drop here...</div>
    <div class="plan-drop-indicator" id="planDropIndicator" style="margin-left:${{labelW}}px;"></div>
  </div>`;

  // Current week shading + indicator line
  const todayTime = new Date().getTime();
  let todayLineHtml = '';
  for (let i = 0; i < weeks.length; i++) {{
    if (todayTime >= weeks[i].start.getTime() && todayTime < weeks[i].end.getTime() + 86400000) {{
      const colLeft = labelW + i * PLAN_CELL;
      todayLineHtml = `<div style="position:absolute;left:${{colLeft}}px;top:0;bottom:0;width:${{PLAN_CELL}}px;background:rgba(46,90,138,0.10);z-index:1;pointer-events:none;"></div>`;
      const frac = (todayTime - weeks[i].start.getTime()) / (7 * 86400000);
      const todayX = labelW + i * PLAN_CELL + frac * PLAN_CELL;
      todayLineHtml += `<div style="position:absolute;left:${{todayX}}px;top:0;bottom:0;width:2px;background:var(--red);z-index:8;pointer-events:none;">
        <div style="position:absolute;top:0;left:-20px;width:42px;text-align:center;background:var(--red);color:white;font-size:9px;font-weight:700;padding:1px 4px;border-radius:0 0 4px 4px;">Today</div>
      </div>`;
      break;
    }}
  }}

  timeline.innerHTML = hdrHtml + `<div style="position:relative;">${{lanesHtml}}${{todayLineHtml}}</div>`;

  // Render backlog sidebar — filtered by product
  const placedNames = new Set(plannedItems.map(p => p.name));
  const filteredBacklog = BACKLOG.filter(b => product === 'All' || b.product === product);
  let blHtml = '';
  filteredBacklog.forEach((b) => {{
    const bi = BACKLOG.indexOf(b);
    const placed = placedNames.has(b.name) ? ' placed' : '';
    blHtml += `<div class="plan-backlog-card${{placed}}" draggable="true" data-bl-idx="${{bi}}">
      <div class="pbc-product">${{esc(b.product)}}</div>
      <div class="pbc-name">${{esc(b.name)}}</div>
    </div>`;
  }});
  backlogList.innerHTML = blHtml;
  const availableCount = filteredBacklog.filter(b => !placedNames.has(b.name)).length;
  backlogCount.textContent = `(${{availableCount}})`;

  // ── Drag from backlog ──
  let dragGhost = null;
  let dragItem = null;

  backlogList.querySelectorAll('.plan-backlog-card:not(.placed)').forEach(card => {{
    card.addEventListener('dragstart', (e) => {{
      const bi = parseInt(card.dataset.blIdx);
      dragItem = BACKLOG[bi];
      card.classList.add('dragging');
      // Create ghost
      dragGhost = document.createElement('div');
      dragGhost.className = 'plan-ghost';
      dragGhost.textContent = dragItem.name;
      document.body.appendChild(dragGhost);
      e.dataTransfer.setDragImage(new Image(), 0, 0); // hide default ghost
      e.dataTransfer.effectAllowed = 'copy';
    }});
    card.addEventListener('dragend', () => {{
      card.classList.remove('dragging');
      if (dragGhost) {{ dragGhost.remove(); dragGhost = null; }}
      dragItem = null;
      const ind = document.getElementById('planDropIndicator');
      if (ind) ind.style.display = 'none';
    }});
  }});

  // Helper: convert fractional week offset to a calendar date
  function offsetToDate(off) {{
    const totalDays = off * 7;
    return new Date(weeks[0].start.getTime() + totalDays * 86400000);
  }}

  timeline.addEventListener('dragover', (e) => {{
    e.preventDefault();
    if (!dragItem) return;
    if (dragGhost) {{
      dragGhost.style.left = (e.clientX + 12) + 'px';
      dragGhost.style.top = (e.clientY + 12) + 'px';
    }}
    // Show drop indicator at fractional position
    const rect = timeline.getBoundingClientRect();
    const x = e.clientX - rect.left + timeline.scrollLeft - labelW;
    const off = Math.max(0, Math.min(weeks.length - 1, x / PLAN_CELL));
    const ind = document.getElementById('planDropIndicator');
    if (ind) {{
      ind.style.display = 'block';
      ind.style.left = (labelW + off * PLAN_CELL) + 'px';
      const indDur = dragItem ? (AVG_DURATION[dragItem.product] || 5) : 5;
      ind.style.width = (indDur * PLAN_CELL) + 'px';
    }}
  }});

  timeline.addEventListener('drop', (e) => {{
    e.preventDefault();
    if (!dragItem) return;
    const rect = timeline.getBoundingClientRect();
    const x = e.clientX - rect.left + timeline.scrollLeft - labelW;
    const defDur = AVG_DURATION[dragItem.product] || 5;
    // Snap to nearest day (1/7 of a week)
    const daySnap = 1 / 7;
    const rawOff = x / PLAN_CELL;
    const off = Math.max(0, Math.min(weeks.length - defDur, Math.round(rawOff / daySnap) * daySnap));
    plannedItems.push({{
      id: Date.now(),
      name: dragItem.name,
      product: dragItem.product,
      startOffset: off,
      duration: defDur,
      startDate: offsetToDate(off).toISOString().slice(0, 10),
    }});
    if (dragGhost) {{ dragGhost.remove(); dragGhost = null; }}
    dragItem = null;
    renderPlan();
    planMarkDirty();
  }});

  // ── Drag to move / resize planned bars ──
  timeline.querySelectorAll('.plan-bar.planned').forEach(bar => {{
    const pi = parseInt(bar.dataset.planIdx);
    const item = plannedItems[pi];

    // Remove button
    bar.querySelector('.plan-remove').addEventListener('mousedown', (e) => {{
      e.stopPropagation();
      plannedItems.splice(pi, 1);
      renderPlan();
      planMarkDirty();
    }});

    // Guide line + date label helpers
    const daySnap = 1 / 7;
    const lanesContainer = timeline.querySelector('div[style*="position:relative"]') || timeline;
    function showGuide(offsetWeeks, edge) {{
      let guide = document.getElementById('planGuide');
      let label = document.getElementById('planGuideLabel');
      if (!guide) {{
        guide = document.createElement('div');
        guide.id = 'planGuide';
        guide.style.cssText = 'position:absolute;top:0;bottom:0;width:1px;border-left:2px dashed var(--blue);z-index:20;pointer-events:none;transition:left 0.05s;';
        lanesContainer.appendChild(guide);
      }}
      if (!label) {{
        label = document.createElement('div');
        label.id = 'planGuideLabel';
        label.style.cssText = 'position:absolute;top:4px;transform:translateX(-50%);background:var(--blue-dark);color:white;font-size:10px;font-weight:600;padding:2px 8px;border-radius:3px;white-space:nowrap;z-index:21;pointer-events:none;box-shadow:0 2px 6px rgba(0,0,0,0.2);';
        lanesContainer.appendChild(label);
      }}
      const px = labelW + offsetWeeks * PLAN_CELL;
      guide.style.left = px + 'px';
      guide.style.display = 'block';
      const d = offsetToDate(offsetWeeks);
      const dayName = d.toLocaleDateString('en-US', {{ weekday: 'short' }});
      label.textContent = dayName + ' ' + d.toLocaleDateString('en-US', {{ month: 'short', day: 'numeric' }});
      label.style.left = px + 'px';
      label.style.display = 'block';
    }}
    function hideGuide() {{
      const g = document.getElementById('planGuide');
      const l = document.getElementById('planGuideLabel');
      if (g) g.style.display = 'none';
      if (l) l.style.display = 'none';
    }}

    // Resize handles
    bar.querySelectorAll('.plan-resize').forEach(handle => {{
      handle.addEventListener('mousedown', (e) => {{
        e.stopPropagation();
        e.preventDefault();
        const edge = handle.dataset.edge;
        const startX = e.clientX;
        const origStart = item.startOffset;
        const origDur = item.duration;

        function onMove(ev) {{
          const dx = ev.clientX - startX;
          const dOff = Math.round((dx / PLAN_CELL) / daySnap) * daySnap;
          if (edge === 'right') {{
            item.duration = Math.max(daySnap, origDur + dOff);
          }} else {{
            const newStart = Math.max(0, origStart + dOff);
            const diff = newStart - origStart;
            item.startOffset = newStart;
            item.duration = Math.max(daySnap, origDur - diff);
          }}
          bar.style.left = (labelW + item.startOffset * PLAN_CELL) + 'px';
          bar.style.width = (item.duration * PLAN_CELL) + 'px';
          // Show guide at the edge being dragged
          const edgeOff = edge === 'right' ? item.startOffset + item.duration : item.startOffset;
          showGuide(edgeOff, edge);
        }}
        function onUp() {{
          document.removeEventListener('mousemove', onMove);
          document.removeEventListener('mouseup', onUp);
          hideGuide();
          item.startDate = offsetToDate(item.startOffset).toISOString().slice(0, 10);
          renderPlan();
          planMarkDirty();
        }}
        document.addEventListener('mousemove', onMove);
        document.addEventListener('mouseup', onUp);
      }});
    }});

    // Drag to move — show guides at both edges
    bar.addEventListener('mousedown', (e) => {{
      if (e.target.classList.contains('plan-resize') || e.target.classList.contains('plan-remove')) return;
      e.preventDefault();
      const startX = e.clientX;
      const origStart = item.startOffset;

      function onMove(ev) {{
        const dx = ev.clientX - startX;
        const dOff = Math.round((dx / PLAN_CELL) / daySnap) * daySnap;
        item.startOffset = Math.max(0, Math.min(weeks.length - item.duration, origStart + dOff));
        bar.style.left = (labelW + item.startOffset * PLAN_CELL) + 'px';
        showGuide(item.startOffset, 'left');
      }}
      function onUp() {{
        document.removeEventListener('mousemove', onMove);
        document.removeEventListener('mouseup', onUp);
        hideGuide();
        item.startDate = offsetToDate(item.startOffset).toISOString().slice(0, 10);
        renderPlan();
        planMarkDirty();
      }}
      document.addEventListener('mousemove', onMove);
      document.addEventListener('mouseup', onUp);
    }});
  }});
}}

function calcBarLeft(start, weeks, labelW) {{
  const t = start.getTime();
  for (let i = 0; i < weeks.length; i++) {{
    if (t <= weeks[i].end.getTime()) {{
      const frac = Math.max(0, (t - weeks[i].start.getTime()) / (7 * 86400000));
      return labelW + i * PLAN_CELL + frac * PLAN_CELL;
    }}
  }}
  return labelW;
}}

function calcBarWidth(start, end, weeks) {{
  const left = calcBarLeft(start, weeks, 0);
  const right = calcBarLeft(end, weeks, 0);
  return Math.max(PLAN_CELL * 0.5, right - left + PLAN_CELL * 0.3);
}}

function planClearAll() {{
  plannedItems = [];
  renderPlan();
  planMarkDirty();
}}

function planExport() {{
  const plan = planBuildPayload();
  const blob = new Blob([JSON.stringify(plan, null, 2)], {{ type: 'application/json' }});
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url; a.download = 'experiment-plan.json'; a.click();
  URL.revokeObjectURL(url);
}}
</script>
</body>
</html>'''


PUBLISH_BRANCH = 'gh-pages'  # the deploy branch GitHub Pages serves from


def deploy_to_github(html_path, repo_path):
    """Copy generated HTML as index.html to the Pages repo's publish branch, commit, and push.

    Publishing lives on its own branch (PUBLISH_BRANCH) so it never collides with the
    source branch (main), which tracks the scripts/data. The browser Planning tab also
    commits plan.json to this branch via the GitHub API, so we sync before adding.
    """
    repo = Path(repo_path).expanduser().resolve()

    if not repo.exists():
        print(f"\n  ERROR: GitHub Pages repo not found at: {repo}")
        print(f"  To set up for the first time:")
        print(f"    1. gh repo create infra-testing-roadmap --public")
        print(f"    2. git clone <repo-url> {repo}")
        print(f"    3. git -C {repo} checkout -b {PUBLISH_BRANCH}")
        print(f"    4. Enable Pages in repo Settings > Pages > Source: {PUBLISH_BRANCH}")
        return False

    if not (repo / '.git').exists():
        print(f"\n  ERROR: {repo} is not a git repository.")
        return False

    # Switch to the publish branch and sync with the remote first, since the browser
    # Planning tab can have committed plan.json there ahead of this local clone.
    try:
        subprocess.run(['git', 'checkout', PUBLISH_BRANCH], cwd=repo, check=True, capture_output=True)
    except subprocess.CalledProcessError as e:
        stderr = e.stderr.decode() if e.stderr else ''
        print(f"\n  ERROR: could not switch to '{PUBLISH_BRANCH}': {stderr or e}")
        print(f"  Create it once with: git -C {repo} checkout -b {PUBLISH_BRANCH}")
        return False
    try:
        subprocess.run(['git', 'fetch', 'origin', PUBLISH_BRANCH], cwd=repo, check=True, capture_output=True, timeout=30)
        subprocess.run(['git', 'rebase', f'origin/{PUBLISH_BRANCH}'], cwd=repo, check=True, capture_output=True, timeout=30)
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired):
        print("  (could not sync with remote first — continuing; push may need a manual pull)")

    # Copy HTML as index.html (skip if already the same file)
    dest = repo / 'index.html'
    src = Path(html_path).resolve()
    if src != dest.resolve():
        shutil.copy2(html_path, dest)
        print(f"  Copied to: {dest}")
    else:
        print(f"  Output is already at: {dest}")

    # Create plan.json if it doesn't exist yet (never clobber browser-saved data)
    plan_file = repo / 'plan.json'
    if not plan_file.exists():
        plan_file.write_text('[]')
        print(f"  Created: {plan_file}")

    # Git add, commit, push
    try:
        subprocess.run(['git', 'add', 'index.html', 'plan.json'], cwd=repo, check=True, capture_output=True)

        result = subprocess.run(
            ['git', 'diff', '--cached', '--quiet'],
            cwd=repo, capture_output=True
        )
        if result.returncode == 0:
            print("  No changes to deploy (HTML is identical to last push).")
            return True

        ts = datetime.now().strftime('%Y-%m-%d %H:%M')
        subprocess.run(
            ['git', 'commit', '-m', f'Update roadmap — {ts}'],
            cwd=repo, check=True, capture_output=True
        )
        print(f"  Committed: Update roadmap — {ts}")

        subprocess.run(
            ['git', 'push', '-u', 'origin', PUBLISH_BRANCH],
            cwd=repo, check=True, capture_output=True, timeout=30
        )
        print("  Pushed to remote.")

        # Try to infer the Pages URL
        remote = subprocess.run(
            ['git', 'remote', 'get-url', 'origin'],
            cwd=repo, capture_output=True, text=True
        )
        if remote.returncode == 0:
            url = remote.stdout.strip()
            # Parse org/repo from SSH or HTTPS URL
            if 'github.com' in url:
                parts = url.rstrip('.git').split('/')
                org, name = parts[-2], parts[-1]
                if ':' in org:
                    org = org.split(':')[-1]
                print(f"\n  Live at: https://{org}.github.io/{name}/")

        return True

    except subprocess.CalledProcessError as e:
        stderr = e.stderr.decode() if e.stderr else ''
        print(f"\n  Git error: {stderr or e}")
        return False
    except subprocess.TimeoutExpired:
        print("\n  Git push timed out. Check your network and try: git push manually.")
        return False


def main():
    parser = argparse.ArgumentParser(
        description='Generate an interactive HTML Gantt from the Infra Testing Excel file.')
    # Default paths point to your OneDrive synced folder
    onedrive_folder = Path.home() / 'Library' / 'CloudStorage' / 'OneDrive-Adobe' / 'Reader & Reduced Mode - Infra Testing'
    parser.add_argument('--input', '-i',
                        default=str(onedrive_folder / 'Reader & Reduced Mode Testing Roadmap.xlsx'),
                        help='Path to the Excel source file (local or OneDrive-synced)')
    parser.add_argument('--output', '-o',
                        default=str(onedrive_folder / 'Product Testing Roadmap.html'),
                        help='Path for the generated HTML file')
    parser.add_argument('--deploy', action='store_true',
                        help='After generating, copy to GitHub Pages repo and push')
    parser.add_argument('--repo',
                        default=str(Path.home() / 'infra-testing-roadmap'),
                        help='Path to the local GitHub Pages repo clone')
    parser.add_argument('--install', action='store_true',
                        help='Copy this script to the OneDrive folder before running')
    parser.add_argument('--serve', action='store_true',
                        help='Start a local server that auto-saves planning data to Excel')
    parser.add_argument('--port', type=int, default=8060,
                        help='Port for the local server (default: 8060)')
    args = parser.parse_args()

    # Self-install: copy this script to OneDrive so it's always up to date
    if args.install:
        script_src = Path(__file__).resolve()
        script_dest = onedrive_folder / 'generate_gantt.py'
        if script_src != script_dest:
            shutil.copy2(script_src, script_dest)
            print(f"  Installed script to: {script_dest}")
        else:
            print(f"  Script already in OneDrive — skipping install.")

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        print(f"\nTip: If the file is on OneDrive/SharePoint, make sure it's synced locally.")
        print(f"     Common paths:")
        print(f"       macOS:   ~/Library/CloudStorage/OneDrive-YourCompany/...")
        print(f"       Windows: C:\\Users\\you\\OneDrive - YourCompany\\...")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  Gantt Generator")
    print(f"{'='*60}")

    data = extract_data(str(input_path))
    html = generate_html(data)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding='utf-8')

    size_kb = output_path.stat().st_size / 1024
    print(f"  Output: {output_path} ({size_kb:.0f} KB)")

    if args.deploy:
        print(f"\n  Deploying to GitHub Pages...")
        success = deploy_to_github(str(output_path), args.repo)
        if not success:
            print(f"\n  Deploy failed. The HTML was still generated at {output_path}")
            sys.exit(1)

    if args.serve:
        print(f"\n  Starting local server on http://localhost:{args.port}")
        print(f"  Planning changes will auto-save to: {input_path}")
        print(f"  Press Ctrl+C to stop.\n")
        start_server(str(output_path), str(input_path), args.port)
    else:
        print(f"{'='*60}")
        print(f"  Done! Open the HTML file in any browser.")
        print(f"  Tip: Run with --serve to enable auto-save from Planning Roadmap.")
        print()


def save_plan_to_excel(excel_path, plan_data):
    """Write planned items to a 'Planning' sheet in the Excel workbook."""
    wb = load_workbook(excel_path)
    # Create or clear the Planning sheet
    if 'Planning' in wb.sheetnames:
        ws = wb['Planning']
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet('Planning')

    # Headers
    headers = ['Product', 'Test Name', 'Planned Start', 'Planned End', 'Duration (Weeks)', 'Fiscal Quarter', 'Fiscal Week']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = cell.font.copy(bold=True)

    # Data rows
    for i, item in enumerate(plan_data, 2):
        ws.cell(i, 1, item.get('product', ''))
        ws.cell(i, 2, item.get('name', ''))
        start_str = item.get('startDate', '')
        end_str = item.get('endDate', '')
        if start_str:
            ws.cell(i, 3, datetime.strptime(start_str, '%Y-%m-%d'))
            ws.cell(i, 3).number_format = 'MMM D, YYYY'
        if end_str:
            ws.cell(i, 4, datetime.strptime(end_str, '%Y-%m-%d'))
            ws.cell(i, 4).number_format = 'MMM D, YYYY'
        ws.cell(i, 5, item.get('durationWeeks', 0))
        fqw = item.get('fqWeek', '')
        ws.cell(i, 6, fqw.split(' W')[0] if ' W' in fqw else fqw)
        ws.cell(i, 7, fqw)

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(30, max(12, max_len + 2))

    wb.save(excel_path)
    return len(plan_data)


def start_server(html_path, excel_path, port):
    """Start a local HTTP server that serves the HTML and handles plan saves."""
    from http.server import HTTPServer, SimpleHTTPRequestHandler
    import json as json_mod

    html_dir = str(Path(html_path).parent)
    html_file = Path(html_path).name

    class Handler(SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=html_dir, **kwargs)

        def do_GET(self):
            if self.path == '/' or self.path == '':
                self.path = '/' + html_file
            super().do_GET()

        def do_POST(self):
            if self.path == '/save-plan':
                content_len = int(self.headers.get('Content-Length', 0))
                body = self.rfile.read(content_len)
                try:
                    plan_data = json_mod.loads(body)
                    count = save_plan_to_excel(excel_path, plan_data)
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json_mod.dumps({'ok': True, 'saved': count}).encode())
                    print(f"  Saved {count} planned items to Excel")
                except Exception as e:
                    self.send_response(500)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json_mod.dumps({'ok': False, 'error': str(e)}).encode())
                    print(f"  Save error: {e}")
            else:
                self.send_response(404)
                self.end_headers()

        def do_OPTIONS(self):
            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
            self.send_header('Access-Control-Allow-Headers', 'Content-Type')
            self.end_headers()

        def log_message(self, fmt, *args):
            # Suppress default GET logging, keep POST logging
            if 'POST' in str(args):
                super().log_message(fmt, *args)

    server = HTTPServer(('localhost', port), Handler)
    try:
        import webbrowser
        webbrowser.open(f'http://localhost:{port}')
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Server stopped.")
        server.server_close()


if __name__ == '__main__':
    main()
