#!/usr/bin/env python3
"""
import_results.py — Reads experiment result screenshots from the `results/` folder,
uses Claude vision to extract test data, cross-references each against the matching
per-experiment .xlsx (by RGS id) to fill in QGNARR/metrics + the full test name, and
appends new rows to the Results sheet in the Testing Roadmap Excel file.

SETUP:
  pip install openpyxl anthropic pillow

USAGE:
  python import_results.py                        # uses default paths (the results/ folder)
  python import_results.py --folder ~/Desktop/results
  python import_results.py --dry-run              # extract + enrich + print, don't write

WORKFLOW:
  1. Drop screenshot images (and the matching RGSxxxx .xlsx files) into the results/ folder
  2. Run this script — it matches each screenshot to its RGS workbook for enrichment
  3. Review the newly appended rows in Excel (highlighted yellow)
  4. Run generate_gantt.py --deploy to regenerate/publish the HTML
  Processed images are renamed to the full test name and moved to results/processed/.
"""

from __future__ import annotations  # allow `X | None` annotations on Python 3.9

import argparse
import base64
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# `anthropic` is imported lazily in main() (only the vision call needs it) so this
# module's Excel/RGS helpers can be reused — e.g. by jira_sync.py — without the package.

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl")
    sys.exit(1)

ONEDRIVE = Path.home() / 'Library' / 'CloudStorage' / 'OneDrive-Adobe' / 'Reader & Reduced Mode - Infra Testing'
DEFAULT_EXCEL  = ONEDRIVE / 'Reader & Reduced Mode Testing Roadmap.xlsx'
DEFAULT_FOLDER = ONEDRIVE / 'results'   # screenshots + per-experiment .xlsx live together here

SUMMARY_TAB_CANDIDATES = ('Final Summary', 'Overall Summary', 'Summary')

IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.webp', '.gif'}

EXTRACT_PROMPT = """You are analyzing a screenshot of an A/B experiment results dashboard for Adobe Acrobat (Reader or Reduced Mode product testing).

Extract ALL experiment results visible in the screenshot. Return a JSON array — one object per experiment row — with these exact keys:

- "product": "Acrobat Reader" or "Acrobat Reduced Mode" (infer from context; default to "Acrobat Reader" if unclear)
- "test_name": full test name including RGS ID, e.g. "RGS0480 - Export PDF Direct Upsell Modal - Reader - Multi - WW - EN"
- "start_date": "YYYY-MM-DD" or null
- "end_date": "YYYY-MM-DD" or null (null if still running)
- "qgnarr": numeric QGNARR value (integer) or null
- "gnarr_lift": decimal lift e.g. 0.022 for 2.2%, or null
- "ctr": decimal CTR lift or null
- "units_pct": decimal units % lift or null
- "winner": one of "Challenger 1", "Challenger 2", "Control", "Rollout", or null
- "details": decision notes or notable commentary as a string, or null
- "gds_url": full URL if a GDS/dashboard link is visible, or null

Return ONLY a valid JSON array with no markdown fences or explanation."""


def encode_image(path: Path) -> tuple[str, str]:
    ext = path.suffix.lower()
    media_map = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                 '.webp': 'image/webp', '.gif': 'image/gif'}
    media_type = media_map.get(ext, 'image/png')
    with open(path, 'rb') as f:
        return base64.standard_b64encode(f.read()).decode(), media_type


def extract_from_screenshot(client: anthropic.Anthropic, image_path: Path) -> list[dict]:
    print(f"  Extracting: {image_path.name}")
    b64, media_type = encode_image(image_path)
    msg = client.messages.create(
        model='claude-opus-4-5',
        max_tokens=2000,
        messages=[{
            'role': 'user',
            'content': [
                {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': b64}},
                {'type': 'text', 'text': EXTRACT_PROMPT}
            ]
        }]
    )
    raw = msg.content[0].text.strip()
    raw = raw.replace('```json', '').replace('```', '').strip()
    parsed = json.loads(raw)
    return parsed if isinstance(parsed, list) else [parsed]


def rgs_id(name) -> str | None:
    """Extract the RGS identifier (e.g. 'RGS0469') from a test name, if present."""
    m = re.search(r'RGS\d+', str(name or ''), re.IGNORECASE)
    return m.group(0).upper() if m else None


def dedup_key(test_name, product):
    """Key a row by (RGS id, product) so the same experiment is caught even when the
    rest of the test name differs, while genuine per-surface variants stay distinct.
    Falls back to the full name when there's no RGS id."""
    rid = rgs_id(test_name)
    prod = str(product or '').strip()
    return (rid, prod) if rid else (None, str(test_name or '').strip())


def get_existing_keys(ws) -> set:
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:  # Test Name is column C (index 2); Product is column B (index 1)
            keys.add(dedup_key(row[2], row[1]))
    return keys


def find_excel_for_rgs(rgs: str | None, results_dir: Path) -> Path | None:
    """Find the per-experiment workbook in results_dir whose filename references this RGS id."""
    if not rgs:
        return None
    for p in sorted(Path(results_dir).glob('*.xlsx')):
        if p.name.startswith('~$'):
            continue          # Excel lock file
        if rgs_id(p.stem) == rgs.upper():
            return p
    return None


def read_excel_summary(excel_path: Path) -> dict:
    """Pull the full test name, date range, and overall metrics from a per-experiment
    workbook's summary tab. Returns a dict with whatever could be found.

    Layout (consistent across the RGS workbooks): a title cell holds
    'RGSxxxx - <name>\\n<MM/DD/YYYY> - <MM/DD/YYYY>\\n~N weeks of baked data', and an
    'Overall - Orders & ARR Scaled' block has rows labelled in col F with
    Control(G) / Challenger1(H) / Lift%(I) / Absolute Lift(J)."""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = next((wb[n] for n in SUMMARY_TAB_CANDIDATES if n in wb.sheetnames), wb[wb.sheetnames[0]])
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    out: dict = {}
    # Title cell — full name (first line) + date range (MM/DD/YYYY - MM/DD/YYYY)
    for row in grid:
        for v in row:
            if isinstance(v, str) and re.search(r'RGS\d+', v):
                lines = [ln.strip() for ln in v.splitlines() if ln.strip()]
                if lines:
                    out['full_name'] = lines[0]
                m = re.search(r'(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})', v)
                if m:
                    out['start_date'], out['end_date'] = m.group(1), m.group(2)
                break
        if 'full_name' in out:
            break

    # Overall metric rows: label@col F (idx5), Control(idx6), Challenger1(idx7), Lift%(idx8)
    for row in grid:
        if len(row) < 9 or not isinstance(row[5], str):
            continue
        label = row[5].strip().lower()
        ch1, lift = row[7], row[8]
        if label == 'units' and isinstance(lift, (int, float)):
            out['units_pct'] = round(float(lift), 4)
        elif label in ('gross new arr', 'gnarr') and isinstance(lift, (int, float)):
            out['gnarr_lift'] = round(float(lift), 4)
            if isinstance(ch1, (int, float)):
                out['qgnarr'] = int(round(ch1))   # Challenger Gross-New-ARR $ (QGNARR mapping — verify)
    return out


def enrich_from_excel(row: dict, results_dir: Path) -> str | None:
    """Cross-reference a screenshot-extracted row against its RGS workbook in results_dir.
    Fills QGNARR/metrics (excel is authoritative over screenshot OCR), upgrades a bare RGS
    id to the full test name, and back-fills dates. Returns the full name (for renaming the
    image) or None if no matching workbook was found."""
    rgs = rgs_id(row.get('test_name'))
    excel = find_excel_for_rgs(rgs, results_dir)
    if not excel:
        return None
    summ = read_excel_summary(excel)
    for k in ('qgnarr', 'gnarr_lift', 'units_pct'):
        if summ.get(k) is not None:
            row[k] = summ[k]
    for k in ('start_date', 'end_date'):
        if not row.get(k) and summ.get(k):
            row[k] = summ[k]
    full = summ.get('full_name')
    if full:
        # upgrade a bare/short test name (e.g. just "RGS0469") to the workbook's full name
        if not row.get('test_name') or len(str(row['test_name'])) < len(full):
            row['test_name'] = full
    return full


def parse_date(val) -> datetime | None:
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except ValueError:
            continue
    return None


def append_row(ws, row: dict, row_num: int):
    # Match column order: A(blank), B(Product), C(Test Name), D(Start), E(End),
    # F(QGNARR), G(GNARR Lift), H(CTR), I(Units %), J(Winner), K(GDS Links),
    # L(Decision/Details), M(Result contrary), N(PM Commentary)
    ws.cell(row_num, 1).value = None
    ws.cell(row_num, 2).value = row.get('product') or ''
    ws.cell(row_num, 3).value = row.get('test_name') or ''
    ws.cell(row_num, 4).value = parse_date(row.get('start_date'))
    ws.cell(row_num, 5).value = parse_date(row.get('end_date'))

    qgnarr = row.get('qgnarr')
    ws.cell(row_num, 6).value = int(qgnarr) if qgnarr is not None else None

    for col, key in [(7, 'gnarr_lift'), (8, 'ctr'), (9, 'units_pct')]:
        v = row.get(key)
        ws.cell(row_num, col).value = float(v) if v is not None else None

    ws.cell(row_num, 10).value = row.get('winner') or None

    gds_url = row.get('gds_url')
    if gds_url:
        cell = ws.cell(row_num, 11)
        cell.value = 'Link'
        cell.hyperlink = gds_url
        cell.font = Font(color='2E75B6', underline='single')
    else:
        ws.cell(row_num, 11).value = None

    ws.cell(row_num, 12).value = row.get('details') or None

    # Date formatting
    for col in (4, 5):
        c = ws.cell(row_num, col)
        if c.value:
            c.number_format = 'YYYY-MM-DD'

    # Highlight new row in light yellow so it's easy to spot
    fill = PatternFill('solid', fgColor='FFFDE7')
    for col in range(1, 15):
        ws.cell(row_num, col).fill = fill


def main():
    parser = argparse.ArgumentParser(description='Import experiment screenshots into the Testing Roadmap Excel.')
    parser.add_argument('--folder', '-f', default=str(DEFAULT_FOLDER),
                        help='Folder containing screenshot images')
    parser.add_argument('--excel', '-e', default=str(DEFAULT_EXCEL),
                        help='Path to the Testing Roadmap Excel file')
    parser.add_argument('--dry-run', action='store_true',
                        help='Extract and print results without writing to Excel')
    parser.add_argument('--api-key', default=os.environ.get('ANTHROPIC_API_KEY'),
                        help='Anthropic API key (or set ANTHROPIC_API_KEY env var)')
    args = parser.parse_args()

    # Resolve paths
    folder = Path(args.folder).expanduser().resolve()
    excel  = Path(args.excel).expanduser().resolve()

    # Validate
    if not folder.exists():
        print(f"Results folder not found: {folder}")
        print(f"Creating it for you...")
        folder.mkdir(parents=True, exist_ok=True)
        print(f"Drop your screenshots and the matching RGSxxxx .xlsx files into:\n  {folder}\nThen run this script again.")
        sys.exit(0)

    images = sorted([p for p in folder.iterdir() if p.suffix.lower() in IMAGE_EXTS])
    if not images:
        print(f"No images found in: {folder}")
        print(f"Supported formats: {', '.join(IMAGE_EXTS)}")
        sys.exit(0)

    print(f"Found {len(images)} image(s) in {folder}")

    if not args.api_key:
        print("\nERROR: Anthropic API key required.")
        print("  Option 1: export ANTHROPIC_API_KEY=sk-ant-...")
        print("  Option 2: python import_results.py --api-key sk-ant-...")
        print("\nGet a key at: https://console.anthropic.com/settings/keys")
        sys.exit(1)

    if not excel.exists():
        print(f"ERROR: Excel file not found: {excel}")
        sys.exit(1)

    try:
        import anthropic
    except ImportError:
        print("ERROR: anthropic package required. Run: pip install anthropic")
        sys.exit(1)
    client = anthropic.Anthropic(api_key=args.api_key)

    # Extract from all screenshots
    all_rows = []
    errors = []
    for img in images:
        try:
            rows = extract_from_screenshot(client, img)
            print(f"    → {len(rows)} row(s) extracted")
            for r in rows:
                r['_source'] = img.name
            all_rows.extend(rows)
        except Exception as e:
            print(f"    ✗ Error: {e}")
            errors.append(img.name)

    if not all_rows:
        print("\nNo data extracted.")
        sys.exit(0)

    # Cross-reference each row against its RGS workbook in the same folder: this fills
    # QGNARR/metrics (excel is authoritative), upgrades a bare RGS id to the full test name,
    # back-fills dates, and tells us what to rename each processed image to.
    rename_to = {}   # source image name -> full test name
    for r in all_rows:
        full = enrich_from_excel(r, folder)
        if full and r.get('_source'):
            rename_to.setdefault(r['_source'], full)
    if rename_to:
        print(f"  Enriched {len(rename_to)} image(s) from matching .xlsx in {folder.name}/")

    print(f"\nTotal: {len(all_rows)} row(s) extracted from {len(images) - len(errors)} image(s)")

    if args.dry_run:
        print("\n── DRY RUN — not writing to Excel ──")
        for r in all_rows:
            print(f"  [{r.get('_source')}] {r.get('test_name')} | {r.get('start_date')} → {r.get('end_date')} | winner: {r.get('winner')} | GNARR: {r.get('gnarr_lift')}")
        sys.exit(0)

    # Write to Excel
    print(f"\nOpening: {excel}")
    wb = load_workbook(excel)
    ws = wb['Results']

    existing = get_existing_keys(ws)
    next_row = ws.max_row + 1
    added = 0
    skipped = 0

    for row in all_rows:
        name = (row.get('test_name') or '').strip()
        key = dedup_key(name, row.get('product'))
        if key in existing:
            print(f"  SKIP (already exists): {name}")
            skipped += 1
            continue
        append_row(ws, row, next_row)
        print(f"  ADDED row {next_row}: {name}")
        existing.add(key)
        next_row += 1
        added += 1

    if added == 0:
        print("\nAll rows already exist in the sheet — nothing to add.")
        sys.exit(0)

    wb.save(excel)
    print(f"\n✓ Saved {added} new row(s) to: {excel}")
    print(f"  (highlighted in yellow — review before regenerating the Gantt)")
    if skipped:
        print(f"  {skipped} duplicate(s) skipped")
    print(f"\nNext step: python generate_gantt.py --deploy")

    # Move processed images to a 'processed' subfolder, renamed to the full test name
    # (from the matched workbook) when known, else left as-is.
    processed_dir = folder / 'processed'
    processed_dir.mkdir(exist_ok=True)
    for img in images:
        if img.name in errors:
            continue
        full = rename_to.get(img.name)
        if full:
            base = re.sub(r'[/\\]', '-', full).strip()
            dest = processed_dir / f'{base}{img.suffix}'
            n = 2
            while dest.exists():            # avoid clobbering multiple shots of one test
                dest = processed_dir / f'{base} ({n}){img.suffix}'
                n += 1
        else:
            dest = processed_dir / img.name
        img.rename(dest)
        print(f"  {img.name}  →  processed/{dest.name}")
    print(f"  Images moved to: {processed_dir}")


if __name__ == '__main__':
    main()
